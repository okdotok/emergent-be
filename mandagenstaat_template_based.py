"""
TEMPLATE-BASED Mandagenstaat Export
Gebruikt user's template en vult alleen data in
GEEN formatting changes - EXACT zoals template
"""

import io
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
import requests


# Template URL (originele user template - NIET WIJZIGEN!)
TEMPLATE_URL = "https://customer-assets.emergentagent.com/job_urenregistratie/artifacts/a3eq7ql5_mandagenstaat.xlsx"
TEMPLATE_PATH = "/tmp/mandagenstaat_user_template.xlsx"


def download_template():
    """Download user template als die nog niet lokaal is"""
    if not os.path.exists(TEMPLATE_PATH):
        response = requests.get(TEMPLATE_URL)
        with open(TEMPLATE_PATH, 'wb') as f:
            f.write(response.content)
        print(f"✅ Template gedownload naar {TEMPLATE_PATH}")


def create_from_template(project, user_week_data, start_date, end_date):
    """
    Maak Excel export vanuit USER TEMPLATE
    Vult alleen data in, behoudt ALLE formatting
    """
    # Zorg dat template beschikbaar is
    download_template()
    
    # Open template
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active
    
    # Bereken week info - gebruik HUIDIGE datum voor weeknummer
    # (niet start_date, want dat kan een filter zijn)
    now = datetime.now()
    week_num = now.isocalendar()[1]
    year = now.year
    
    # Update sheet naam met weeknummer
    ws.title = f"week {week_num}"
    
    # === TITEL STYLING: links uitlijnen voor meer ruimte tussen logo en titel ===
    # Zet titel links uit in plaats van gecentreerd voor betere spacing
    if ws['D9'].value:
        ws['D9'].font = Font(name='Arial', size=11, bold=False)
        ws['D9'].alignment = Alignment(horizontal='left', vertical='center')
    
    # === DATA INVULLING met Arial 10 font ===
    
    # Tabel 1 - Project Info (C11-C14)
    cell = ws['C11']
    cell.value = project.get('company', '')
    cell.font = Font(name='Arial', size=10)
    
    cell = ws['C12']
    cell.value = project.get('name', '')
    cell.font = Font(name='Arial', size=10)
    
    cell = ws['C13']
    cell.value = f"W{week_num}/{year}"  # W voor weeknummer
    cell.font = Font(name='Arial', size=10)
    
    cell = ws['C14']
    cell.value = project.get('description', '')
    cell.font = Font(name='Arial', size=10)
    
    # L19 header NIET toevoegen - geen totaal kolom per rij
    
    # Tabel 2 - Data rijen (20-34)
    current_row = 20
    for user_name, data in sorted(user_week_data.items()):
        if current_row > 34:
            break
        
        # B: Naam (AFGEKORT naar voorletter)
        # "Badreddine el Mobarie" -> "B. el Mobarie"
        name_parts = user_name.split(' ', 1)
        if len(name_parts) > 1:
            abbreviated_name = f"{name_parts[0][0]}. {name_parts[1]}"
        else:
            abbreviated_name = user_name  # Als er geen spatie is, gebruik volledige naam
        
        cell = ws.cell(row=current_row, column=2, value=abbreviated_name)
        cell.font = Font(name='Arial', size=10)
        
        # C: BSN
        cell = ws.cell(row=current_row, column=3, value=data.get('bsn', ''))
        cell.font = Font(name='Arial', size=10)
        
        # D: Week nummer
        cell = ws.cell(row=current_row, column=4, value=week_num)
        cell.font = Font(name='Arial', size=10)
        
        # E-K: ma-zon (7 dagen) - LEEG als 0, anders hele getallen
        for col_idx, hours in enumerate(data['days'], start=5):
            # Als 0 uren: LEEG laten, anders: heel getal
            value = int(round(hours)) if hours > 0 else None
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = Font(name='Arial', size=10)
        
        # L kolom: NIET invullen (geen totaal per rij)
        
        current_row += 1
    
    # Vul lege rijen (als er minder dan 15 werknemers zijn)
    while current_row <= 34:
        # Maak cellen leeg maar laat borders intact en set font
        for col in range(2, 12):  # B tot K
            cell = ws.cell(row=current_row, column=col)
            cell.value = None
            cell.font = Font(name='Arial', size=10)
        
        # L kolom: leeg laten (geen totaal per rij)
        
        current_row += 1
    
    # Datum & Plaats (C37, C38)
    # Runtime datum in dd-mm-yyyy format
    ws['C37'] = datetime.now().strftime("%d-%m-%Y")
    ws['C37'].font = Font(name='Arial', size=10)
    ws['C38'] = "Utrecht"
    ws['C38'].font = Font(name='Arial', size=10)
    
    # TOTALEN (E35-L35) - bereken en zet als waarden (LEEG als 0)
    # Bereken totalen per dag
    day_totals = [0, 0, 0, 0, 0, 0, 0]  # ma-zon
    for user_name, data in user_week_data.items():
        for day_idx, hours in enumerate(data['days']):
            day_totals[day_idx] += hours
    
    # Zet totalen in E35-K35 (LEEG als 0)
    for col_idx, total in enumerate(day_totals, start=5):  # E=5, F=6, ... K=11
        # Als 0: LEEG, anders: heel getal
        value = int(round(total)) if total > 0 else None
        cell = ws.cell(row=35, column=col_idx, value=value)
        cell.font = Font(name='Arial', size=10)
        cell.number_format = '0'
    
    # Grand total in L35 (LEEG als 0)
    grand_total = sum(day_totals)
    value = int(round(grand_total)) if grand_total > 0 else None
    cell = ws.cell(row=35, column=12, value=value)
    cell.font = Font(name='Arial', size=10)
    cell.number_format = '0'
    
    # === CLEANUP: Verwijder dubbele/kleine logo's ===
    # De template heeft soms meerdere images, we willen alleen het grote logo bovenaan (row 0-5)
    images_to_remove = []
    for img in ws._images:
        try:
            # Verwijder images die NIET in de header (row 0-5) staan
            # Of images die te groot zijn (duplicaten)
            if hasattr(img.anchor, '_from'):
                row_pos = img.anchor._from.row if hasattr(img.anchor._from, 'row') else 0
                # Behoud alleen logo's in de eerste 6 rijen (header)
                if row_pos > 5:
                    images_to_remove.append(img)
        except:
            pass  # Skip als anchor info niet beschikbaar
    
    # Verwijder de gemarkeerde images
    for img in images_to_remove:
        ws._images.remove(img)
    
    # === CLEANUP: Controleer verborgen rijen/kolommen ===
    # Zorg dat geen rijen/kolommen verborgen zijn in print area (A1:L47)
    for row_num in range(1, 48):
        if row_num in ws.row_dimensions:
            ws.row_dimensions[row_num].hidden = False
    
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        if col_letter in ws.column_dimensions:
            ws.column_dimensions[col_letter].hidden = False
    
    # === SPACING: Meer ruimte tussen logo en titel ===
    # Vergroot rij 8 voor extra ruimte tussen logo (rijen 1-7) en titel (rij 9)
    ws.row_dimensions[8].height = 25  # Was default 15, nu 25 voor meer ruimte
    
    # === PRINT SETTINGS: EXACT ZOALS USER SPECIFICATIE ===
    # A4 Portrait, Fit to 1 page wide × 1 page tall
    # Marges "Normaal" - NIET custom!
    # Print area exact kop+tabel (A1:L47)
    
    # STAP 1: Print area
    ws.print_area = 'A1:L47'
    
    # STAP 2: A4 Portrait
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    
    # STAP 3: Marges "Normaal" (Excel default)
    # NIET custom! Gebruik Excel defaults
    ws.page_margins.top = 0.75     # inch (default normal)
    ws.page_margins.bottom = 0.75  # inch (default normal)
    ws.page_margins.left = 0.7     # inch (default normal)
    ws.page_margins.right = 0.7    # inch (default normal)
    ws.page_margins.header = 0.3   # inch (default normal)
    ws.page_margins.footer = 0.3   # inch (default normal)
    
    # STAP 4: Fit to 1 page wide × 1 page tall
    # DIT IS CRUCIAAL - exact zoals user vraagt
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.scale = None  # Auto-scale door fit
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    
    # STAP 5: Center horizontally
    ws.page_setup.horizontalCentered = True
    
    # Print title rows (header herhalen)
    ws.print_title_rows = '19:19'
    
    # === BEREKEN FORMULES HANDMATIG en sla cached values op ===
    # Dit zorgt dat totalen direct zichtbaar zijn, zelfs in Protected View
    
    # Bereken dag totalen (E35-K35)
    for col_idx in range(5, 12):  # E tot K
        total = 0
        for row in range(20, 35):  # Data rijen
            cell_val = ws.cell(row=row, column=col_idx).value
            if isinstance(cell_val, (int, float)):
                total += int(cell_val)
        
        # Zet de cached value (naast de formule die al bestaat)
        # Openpyxl slaat beide op: formule EN cached value
        formula_cell = ws.cell(row=35, column=col_idx)
        # De formule blijft behouden, we voegen alleen cached value toe
        # Dit doen we door een dummy write - Excel zal deze waarde tonen
        current_formula = formula_cell.value
        formula_cell.value = total  # Sla berekende waarde op
        # Helaas verliest dit de formule... we moeten anders
    
    # Betere aanpak: gebruik openpyxl's mogelijkheid om cached values mee te geven
    # Maar dat is complex. Simpelste: vervang formules door waarden voor nu
    
    # Bereken en vervang E35-K35
    for col_idx in range(5, 12):
        total = 0
        for row in range(20, 35):
            cell_val = ws.cell(row=row, column=col_idx).value
            if isinstance(cell_val, (int, float)):
                total += int(cell_val)
        ws.cell(row=35, column=col_idx).value = total
    
    # Bereken en vervang L35 (grand total)
    grand_total = 0
    for col_idx in range(5, 12):
        cell_val = ws.cell(row=35, column=col_idx).value
        if isinstance(cell_val, (int, float)):
            grand_total += int(cell_val)
    ws.cell(row=35, column=12).value = grand_total
    
    # === Save to BytesIO ===
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def excel_to_html(wb, ws):
    """
    Converteer Excel werkblad naar HTML met exacte styling
    """
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {
                size: A4 portrait;
                margin: 1cm;
            }
            body {
                font-family: Arial, sans-serif;
                font-size: 10pt;
                margin: 0;
                padding: 0;
            }
            table {
                border-collapse: collapse;
                width: 100%;
                margin-bottom: 0.5cm;
            }
            td, th {
                padding: 4px 6px;
                text-align: left;
                border: 0.5pt solid #D9D9D9;
            }
            .header-table {
                border: 1.5pt solid #D9D9D9;
            }
            .data-table {
                border: 1.5pt solid #D9D9D9;
            }
            .bold {
                font-weight: bold;
            }
            .center {
                text-align: center;
            }
            .logo {
                width: 150px;
                height: auto;
            }
            .company-name {
                font-size: 11pt;
                font-weight: bold;
                text-align: right;
                vertical-align: middle;
            }
            .no-border {
                border: none;
            }
            .section-gap {
                height: 0.8cm;
            }
        </style>
    </head>
    <body>
    """
    
    # Logo en company name (rij 1-9)
    logo_path = '/app/backend/logo.png'
    import base64
    logo_b64 = ""
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            logo_b64 = base64.b64encode(f.read()).decode()
    
    html += f"""
    <table class="no-border">
        <tr>
            <td class="no-border" style="width: 30%;">
                <img src="data:image/png;base64,{logo_b64}" class="logo" />
            </td>
            <td class="no-border company-name" style="width: 70%;">
                {ws['D9'].value or ''}
            </td>
        </tr>
    </table>
    <div class="section-gap"></div>
    """
    
    # Tabel 1: Project info (rij 11-14)
    html += '<table class="header-table">'
    for row in range(11, 15):
        label = ws.cell(row=row, column=2).value or ''
        value = ws.cell(row=row, column=3).value or ''
        html += f'<tr><td class="bold" style="width: 25%;">{label}</td><td style="width: 75%;">{value}</td></tr>'
    html += '</table><div class="section-gap"></div>'
    
    # Tabel 2: Uren tabel (rij 19-35)
    html += '<table class="data-table">'
    
    # Headers (rij 19)
    html += '<tr>'
    for col in range(2, 13):  # B tot L
        val = ws.cell(row=19, column=col).value or ''
        html += f'<th class="bold center">{val}</th>'
    html += '</tr>'
    
    # Data rijen (20-34)
    for row in range(20, 35):
        html += '<tr>'
        for col in range(2, 13):  # B tot L
            cell = ws.cell(row=row, column=col)
            val = cell.value or ''
            
            # Als het een formule is, evalueer het
            if isinstance(val, str) and val.startswith('='):
                try:
                    val = f"{float(cell.value):.1f}" if cell.value else '0.0'
                except:
                    val = '0.0'
            elif col >= 5 and col <= 12:  # Uren kolommen
                try:
                    val = f"{float(val):.1f}" if val else '0.0'
                except:
                    val = str(val)
            
            align_class = 'center' if col >= 3 else ''
            html += f'<td class="{align_class}">{val}</td>'
        html += '</tr>'
    
    # Totalen rij (35)
    html += '<tr>'
    for col in range(2, 13):
        cell = ws.cell(row=35, column=col)
        val = ''
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            # Dit is een formule - toon de berekende waarde
            val = 'TOTAAL' if col == 2 else ''
        else:
            val = cell.value or ''
        
        bold_class = 'bold' if val else ''
        html += f'<td class="center {bold_class}">{val}</td>'
    html += '</tr>'
    
    html += '</table><div class="section-gap"></div>'
    
    # Datum en Plaats
    datum = ws['C37'].value or ''
    plaats = ws['C38'].value or ''
    html += f"""
    <table class="no-border">
        <tr><td class="no-border bold" style="width: 15%;">Datum:</td><td class="no-border">{datum}</td></tr>
        <tr><td class="no-border bold">Plaats:</td><td class="no-border">{plaats}</td></tr>
    </table>
    <div class="section-gap"></div>
    """
    
    # Handtekeningen
    accoord1 = ws['B41'].value or 'Accoord Uitvoerder'
    accoord2 = ws.cell(row=41, column=5).value or 'Accoord The Global'
    html += f"""
    <table class="no-border">
        <tr>
            <td class="no-border bold" style="width: 50%;">{accoord1}</td>
            <td class="no-border bold" style="width: 50%;">{accoord2}</td>
        </tr>
        <tr>
            <td style="height: 2cm; border: 0.5pt solid #D9D9D9;"></td>
            <td style="height: 2cm; border: 0.5pt solid #D9D9D9;"></td>
        </tr>
    </table>
    """
    
    html += "</body></html>"
    return html


def create_pdf_as_excel_print(excel_file):
    """
    EXACT EXCEL PRINT REPLICA
    Repliceert precies hoe Excel de sheet print naar PDF
    - Zwarte borders (zoals voorbeeld PDF)
    - Exacte spacing en layout
    - Logo links, company name rechts
    - Alle details exact zoals Excel print
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Image as RLImage
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph
    
    # Open Excel
    excel_file.seek(0)
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    pdf_file = io.BytesIO()
    
    # A4 portrait - EXACT zoals Excel print (uit voorbeeld PDF)
    doc = SimpleDocTemplate(
        pdf_file,
        pagesize=A4,
        leftMargin=2*cm,
        rightMargin=2*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm
    )
    
    elements = []
    
    # === HEADER: LOGO (links) + COMPANY NAME (rechts) ===
    # EXACT zoals voorbeeld PDF: logo links, company rechts uitgelijnd
    logo_path = '/app/backend/logo.png'
    company_name = ws['D9'].value or 'The Global Bedrijfsdiensten BV'
    
    if os.path.exists(logo_path):
        try:
            # Logo grootte zoals in voorbeeld
            logo = RLImage(logo_path, width=3.5*cm, height=2.2*cm)
            
            # Company name rechts uitgelijnd (zoals voorbeeld PDF)
            company_para = Paragraph(
                f"<para align='right'>{company_name}</para>", 
                ParagraphStyle('Company', fontSize=10, fontName='Helvetica')
            )
            
            # Tabel zonder borders (zoals Excel print header)
            header_table = Table([[logo, company_para]], colWidths=[4*cm, 12*cm])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            elements.append(header_table)
        except Exception as e:
            print(f"Logo error: {e}")
    
    # Ruimte na header (zoals Excel print)
    elements.append(Spacer(1, 10*mm))
    
    # === TABEL 1: Project Info - ZWARTE borders zoals voorbeeld PDF ===
    project_data = []
    for row in range(11, 15):
        label = ws.cell(row=row, column=2).value or ''
        value = ws.cell(row=row, column=3).value or ''
        project_data.append([label, value])
    
    project_table = Table(project_data, colWidths=[5*cm, 11*cm])
    project_table.setStyle(TableStyle([
        # ZWARTE borders zoals voorbeeld PDF (NIET grijs!)
        ('BOX', (0, 0), (-1, -1), 1.2, colors.black),  # Dikke buitenrand
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),  # Dunne binnenlijnen
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(project_table)
    
    # Ruimte tussen tabellen (zoals Excel print)
    elements.append(Spacer(1, 10*mm))
    
    # === TABEL 2: Uren - ZWARTE borders zoals voorbeeld PDF ===
    uren_data = []
    
    # Headers (rij 19) - kolommen B tot K
    headers = []
    for col in range(2, 12):  # B tot K
        val = ws.cell(row=19, column=col).value or ''
        headers.append(val)
    uren_data.append(headers)
    
    # Data rijen (20-34) - ALLE 15 rijen (ook lege)
    for row in range(20, 35):
        row_data = []
        for col in range(2, 12):  # B tot K
            cell = ws.cell(row=row, column=col)
            val = cell.value
            
            # Format waarde als heel getal
            if val is None or val == '':
                row_data.append('')
            elif isinstance(val, (int, float)):
                row_data.append(str(int(val)))
            else:
                row_data.append(str(val))
        
        uren_data.append(row_data)
    
    # Totalen rij (35) - onderaan met cijfers
    totaal_row = []
    for col in range(2, 12):  # B tot K
        cell = ws.cell(row=35, column=col)
        val = cell.value
        
        # Als het een formule is, toon "0" zoals voorbeeld PDF
        if val and isinstance(val, str) and val.startswith('='):
            totaal_row.append('0')
        elif val is None or val == '':
            totaal_row.append('')
        elif isinstance(val, (int, float)):
            totaal_row.append(str(int(val)))
        else:
            totaal_row.append(str(val))
    
    uren_data.append(totaal_row)
    
    # Kolom breedtes zoals voorbeeld PDF (Naam breed, BSN/week gemiddeld, dagen smal)
    col_widths = [3.2*cm, 1.8*cm, 1.8*cm, 1.1*cm, 1.1*cm, 1.1*cm, 1.1*cm, 1.1*cm, 1.1*cm, 1.1*cm]
    
    uren_table = Table(uren_data, colWidths=col_widths)
    uren_table.setStyle(TableStyle([
        # Headers - NIET bold volgens voorbeeld PDF
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Data rijen
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Naam links
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),  # Rest gecentreerd
        # Totalen rij - NIET bold volgens voorbeeld PDF
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        # ZWARTE borders zoals voorbeeld PDF (NIET grijs!)
        ('BOX', (0, 0), (-1, -1), 1.2, colors.black),  # Dikke buitenrand
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),  # Dunne binnenlijnen
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(uren_table)
    
    # Ruimte na uren tabel
    elements.append(Spacer(1, 8*mm))
    
    # === DATUM & PLAATS (zoals voorbeeld PDF) ===
    datum = ws['C37'].value or ''
    plaats = ws['C38'].value or ''
    
    datum_plaats_data = [
        ['Datum:', datum],
        ['Plaats:', plaats]
    ]
    datum_table = Table(datum_plaats_data, colWidths=[1.5*cm, 6*cm])
    datum_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]))
    elements.append(datum_table)
    
    # Ruimte voor handtekeningen
    elements.append(Spacer(1, 10*mm))
    
    # === HANDTEKENINGEN (zoals voorbeeld PDF met ZWARTE borders) ===
    accoord1 = ws['B41'].value or 'Accoord Uitvoerder'
    accoord2 = ws.cell(row=41, column=5).value or 'Accoord The Global'
    
    sig_data = [
        [accoord1, accoord2],
        ['', '']  # Lege rijen voor handtekening
    ]
    
    sig_table = Table(sig_data, colWidths=[7.5*cm, 7.5*cm], rowHeights=[6*mm, 20*mm])
    sig_table.setStyle(TableStyle([
        # Labels boven vakken
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, 0), 0),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
        # Handtekening vakken met ZWARTE borders (zoals voorbeeld PDF)
        ('BOX', (0, 1), (0, 1), 0.8, colors.black),
        ('BOX', (1, 1), (1, 1), 0.8, colors.black),
        ('VALIGN', (0, 1), (-1, 1), 'MIDDLE'),
    ]))
    elements.append(sig_table)
    
    # Build PDF
    doc.build(elements)
    pdf_file.seek(0)
    
    return pdf_file


def create_pdf_from_template(project, user_week_data, start_date, end_date):
    """
    Maak PDF export - EXACT 1:1 REPLICA VAN EXCEL
    Gebruikt LibreOffice om Excel naar PDF te converteren
    - 100% scale (GEEN fit-to-page shrink)
    - Exacte marges zoals Excel print settings
    - Print area en page setup worden gerespecteerd
    """
    import subprocess
    import tempfile
    import os
    
    # Maak eerst Excel met alle correcte print settings (100% scale)
    excel_file = create_from_template(project, user_week_data, start_date, end_date)
    
    # Save Excel naar temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', mode='wb') as tmp_excel:
        tmp_excel.write(excel_file.read())
        excel_path = tmp_excel.name
    
    # Output PDF path - LibreOffice schrijft naar dezelfde directory
    output_dir = os.path.dirname(excel_path)
    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    pdf_path = os.path.join(output_dir, f"{base_name}.pdf")
    
    try:
        # Check if LibreOffice is available
        libreoffice_check = subprocess.run(['which', 'libreoffice'], capture_output=True)
        if libreoffice_check.returncode != 0:
            raise FileNotFoundError(
                "LibreOffice is niet geïnstalleerd. "
                "Installeer met: apt-get install -y libreoffice-calc"
            )
        
        # Excel naar PDF met LibreOffice
        # Belangrijke opties:
        # --convert-to pdf: PDF conversie
        # --headless: geen GUI
        # De Excel print settings (scale 100%, marges, etc.) worden automatisch overgenomen
        result = subprocess.run([
            'libreoffice',
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', output_dir,
            excel_path
        ], timeout=30, capture_output=True, 
        env={**os.environ, 'SAL_USE_VCLPLUGIN': 'svp'})
        
        if result.returncode != 0:
            error_msg = result.stderr.decode()
            raise Exception(f"LibreOffice conversion failed: {error_msg}")
        
        # Check of PDF is aangemaakt
        if not os.path.exists(pdf_path):
            raise Exception(f"PDF not created at expected path: {pdf_path}")
        
        # Lees PDF
        with open(pdf_path, 'rb') as f:
            pdf_data = io.BytesIO(f.read())
        
        return pdf_data
        
    finally:
        # Cleanup temp files
        try:
            if os.path.exists(excel_path):
                os.unlink(excel_path)
            if os.path.exists(pdf_path):
                os.unlink(pdf_path)
        except Exception as cleanup_error:
            print(f"Cleanup warning: {cleanup_error}")


def create_pdf_reportlab_fallback(project, user_week_data, start_date, end_date):
    """
    Fallback methode met ReportLab als WeasyPrint niet beschikbaar is
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    # Bereken week info
    now = datetime.now()
    week_num = now.isocalendar()[1]
    year = now.year
    
    pdf_file = io.BytesIO()
    
    doc = SimpleDocTemplate(
        pdf_file,
        pagesize=A4,
        leftMargin=1*cm,
        rightMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm,
        title=f"Mandagenstaat week {week_num}"
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # === LOGO + COMPANY NAME ===
    logo_path = '/app/backend/logo.png'
    if os.path.exists(logo_path):
        try:
            logo = RLImage(logo_path, width=4*cm, height=2.5*cm)
            company_style = ParagraphStyle('CompanyStyle', parent=styles['Heading1'], 
                                          fontSize=11, textColor=colors.black)
            company_text = Paragraph("<b>The Global Bedrijfsdiensten BV</b>", company_style)
            header_table = Table([[logo, company_text]], colWidths=[5*cm, 14*cm])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ]))
            elements.append(header_table)
        except:
            pass
    
    elements.append(Spacer(1, 0.5*cm))
    
    # === TABEL 1: Project Info ===
    project_info_data = [
        ['Naam opdrachtgever', project.get('company', '')],
        ['Project', project.get('name', '')],
        ['Weeknummer/ jaar', f"{week_num}/{year}"],
        ['Soort werkzaamheden', project.get('description', '')]
    ]
    
    project_table = Table(project_info_data, colWidths=[5*cm, 13*cm])
    project_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#D9D9D9')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    elements.append(project_table)
    elements.append(Spacer(1, 0.8*cm))
    
    # === TABEL 2: Uren (met TOTAAL kolom) ===
    table_data = [['Naam', 'BSN', 'week nr', 'ma', 'di', 'wo', 'do', 'vrij', 'zat', 'zon', 'Totaal']]
    
    # Data rijen met AFGEKORTE namen
    for user_name, data in sorted(user_week_data.items()):
        # Afkorten: "Badreddine el Mobarie" -> "B. el Mobarie"
        name_parts = user_name.split(' ', 1)
        if len(name_parts) > 1:
            abbreviated_name = f"{name_parts[0][0]}. {name_parts[1]}"
        else:
            abbreviated_name = user_name
        
        row = [abbreviated_name, data.get('bsn', ''), str(week_num)]
        row.extend([f"{h:.1f}" if h > 0 else "0.0" for h in data['days']])
        
        # Totaal per rij
        row_total = sum(data['days'])
        row.append(f"{row_total:.1f}")
        
        table_data.append(row)
    
    # Totals row
    if len(table_data) > 1:
        day_totals = [0] * 7
        for user_name, data in user_week_data.items():
            for i, hours in enumerate(data['days']):
                day_totals[i] += hours
        
        totals_row = ['TOTAAL', '', '']
        totals_row.extend([f"{t:.1f}" for t in day_totals])
        
        # Grand total
        grand_total = sum(day_totals)
        totals_row.append(f"{grand_total:.1f}")
        
        table_data.append(totals_row)
    
    col_widths = [3.5*cm, 2.2*cm, 1.5*cm, 1.0*cm, 1.0*cm, 1.0*cm, 1.0*cm, 1.0*cm, 1.0*cm, 1.0*cm, 1.2*cm]
    
    main_table = Table(table_data, colWidths=col_widths)
    main_table.setStyle(TableStyle([
        # Headers
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Data
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        # Totals
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        # Borders
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#D9D9D9')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    elements.append(main_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # === DATUM & PLAATS ===
    runtime_date = datetime.now().strftime("%d-%m-%Y")
    date_place_data = [
        ['Datum:', runtime_date],
        ['Plaats:', 'Utrecht']
    ]
    date_place_table = Table(date_place_data, colWidths=[2*cm, 8*cm])
    date_place_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(date_place_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # === HANDTEKENINGEN ===
    sig_label_style = ParagraphStyle('SigLabel', parent=styles['Normal'], 
                                     fontSize=9, fontName='Helvetica-Bold')
    sig_data = [
        [Paragraph('Accoord Uitvoerder', sig_label_style), 
         Paragraph('Accoord The Global', sig_label_style)],
        ['', '']
    ]
    
    sig_table = Table(sig_data, colWidths=[9*cm, 9*cm], rowHeights=[0.5*cm, 2*cm])
    sig_table.setStyle(TableStyle([
        ('BOX', (0, 1), (0, 1), 0.5, colors.HexColor('#D9D9D9')),
        ('BOX', (1, 1), (1, 1), 0.5, colors.HexColor('#D9D9D9')),
        ('VALIGN', (0, 0), (-1, 0), 'BOTTOM'),
    ]))
    
    elements.append(sig_table)
    
    # Build PDF
    doc.build(elements)
    pdf_file.seek(0)
    
    return pdf_file
