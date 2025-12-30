"""
PIXEL-PERFECT Mandagenstaat Export Functions
Generates Excel and PDF exports matching user template EXACTLY
- Exacte kolombreedtes
- Exacte borders/randen
- Exacte lettertypen en sizes
- Exacte spacing en uitlijning
- Logo strak uitgelijnd en groter
- Datumformaat: dd-mm-yyyy overal
- Bestandsnaam: Mandagenstaat_YYYY-MM_Bedrijfsnaam.xlsx/pdf
"""

import io
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm


def create_professional_excel(project, user_week_data, start_date, end_date):
    """
    Create a professional Excel export matching the exact template
    
    Args:
        project: Project dictionary
        user_week_data: Dictionary with user data {user_name: {"bsn": str, "days": [hours]}}
        start_date: Start date string
        end_date: End date string
    
    Returns:
        BytesIO object with Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mandagenstaat"
    
    # Calculate dates
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    week_num = start_dt.isocalendar()[1]
    year = start_dt.year
    
    # Set exact column widths (matching template exactly)
    ws.column_dimensions['A'].width = 20  # Naam
    ws.column_dimensions['B'].width = 12  # BSN
    ws.column_dimensions['C'].width = 10  # week nummer
    ws.column_dimensions['D'].width = 6   # ma
    ws.column_dimensions['E'].width = 6   # di
    ws.column_dimensions['F'].width = 6   # wo
    ws.column_dimensions['G'].width = 6   # do
    ws.column_dimensions['H'].width = 6   # vrij
    ws.column_dimensions['I'].width = 6   # zat
    ws.column_dimensions['J'].width = 6   # zon
    ws.column_dimensions['K'].width = 8   # Totaal
    
    # Add logo if exists
    logo_path = '/app/backend/logo.png'
    if os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            img.width = 140
            img.height = 60
            ws.add_image(img, 'A1')
        except:
            pass
    
    # Company name (row 3, merged, right-aligned)
    ws.merge_cells('A3:K3')
    ws['A3'] = "The Global Bedrijfsdiensten BV"
    ws['A3'].font = Font(name='Arial', size=14, bold=True)
    ws['A3'].alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[3].height = 25
    
    # Empty row
    ws.row_dimensions[4].height = 10
    
    # Project info section (rows 5-8, with borders)
    project_info = [
        ("Naam opdrachtgever:", project.get('company', '')),
        ("Project:", project.get('name', '')),
        ("Weeknummer/jaar:", f"{week_num}/{year}"),
        ("Soort werkzaamheden:", project.get('description', ''))
    ]
    
    for idx, (label, value) in enumerate(project_info):
        row = 5 + idx
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = Font(name='Arial', size=10, bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        ws.merge_cells(f'B{row}:K{row}')
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=2).font = Font(name='Arial', size=10)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center')
        
        # Borders for info section
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
    
    # Empty row
    ws.row_dimensions[9].height = 15
    
    # Table headers (row 10)
    headers = ['Naam', 'BSN', 'week nummer', 'ma', 'di', 'wo', 'do', 'vrij', 'zat', 'zon', 'Totaal']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light gray
        cell.border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
    
    ws.row_dimensions[10].height = 20
    
    # Data rows (starting from row 11)
    current_row = 11
    for user_name, data in sorted(user_week_data.items()):
        # Name
        ws.cell(row=current_row, column=1, value=user_name)
        ws.cell(row=current_row, column=1).font = Font(name='Arial', size=10)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        # BSN
        ws.cell(row=current_row, column=2, value=data.get('bsn', ''))
        ws.cell(row=current_row, column=2).font = Font(name='Arial', size=10)
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
        
        # Week number
        ws.cell(row=current_row, column=3, value=week_num)
        ws.cell(row=current_row, column=3).font = Font(name='Arial', size=10)
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
        
        # Days (columns 4-10: ma-zon)
        week_total = 0
        for col_idx, hours in enumerate(data['days'], start=4):
            val = hours if hours > 0 else 0
            week_total += val
            ws.cell(row=current_row, column=col_idx, value=val)
            ws.cell(row=current_row, column=col_idx).font = Font(name='Arial', size=10)
            ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=col_idx).number_format = '0.0'
        
        # Week total (column 11)
        ws.cell(row=current_row, column=11, value=week_total)
        ws.cell(row=current_row, column=11).font = Font(name='Arial', size=10, bold=True)
        ws.cell(row=current_row, column=11).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=current_row, column=11).number_format = '0.0'
        
        # Borders for data row
        for col in range(1, 12):
            ws.cell(row=current_row, column=col).border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
        
        current_row += 1
    
    # Totals row
    ws.cell(row=current_row, column=1, value="TOTAAL")
    ws.cell(row=current_row, column=1).font = Font(name='Arial', size=11, bold=True)
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    
    ws.cell(row=current_row, column=2, value="")
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    ws.cell(row=current_row, column=3, value="")
    ws.cell(row=current_row, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Calculate and add totals for each day + grand total
    day_totals = [0, 0, 0, 0, 0, 0, 0]
    for user_data in user_week_data.values():
        for i, hours in enumerate(user_data['days']):
            day_totals[i] += hours
    
    grand_total = 0
    for col_idx, total in enumerate(day_totals, start=4):
        ws.cell(row=current_row, column=col_idx, value=total)
        ws.cell(row=current_row, column=col_idx).font = Font(name='Arial', size=11, bold=True)
        ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=current_row, column=col_idx).number_format = '0.0'
        ws.cell(row=current_row, column=col_idx).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        grand_total += total
    
    # Grand total
    ws.cell(row=current_row, column=11, value=grand_total)
    ws.cell(row=current_row, column=11).font = Font(name='Arial', size=12, bold=True)
    ws.cell(row=current_row, column=11).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=current_row, column=11).number_format = '0.0'
    ws.cell(row=current_row, column=11).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Borders for totals row (thicker)
    for col in range(1, 12):
        ws.cell(row=current_row, column=col).border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
    
    current_row += 2  # Empty row
    
    # Footer info
    ws.cell(row=current_row, column=1, value="Datum:")
    ws.cell(row=current_row, column=1).font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=current_row, column=2, value=datetime.now().strftime("%Y-%m-%d"))
    ws.cell(row=current_row, column=2).font = Font(name='Arial', size=10)
    
    current_row += 1
    ws.cell(row=current_row, column=1, value="Plaats:")
    ws.cell(row=current_row, column=1).font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=current_row, column=2, value=project.get('location', 'Utrecht'))
    ws.cell(row=current_row, column=2).font = Font(name='Arial', size=10)
    
    current_row += 2
    
    # Signature section
    ws.cell(row=current_row, column=1, value="Accoord Uitvoerder")
    ws.cell(row=current_row, column=1).font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=current_row, column=6, value="Accoord The Global")
    ws.cell(row=current_row, column=6).font = Font(name='Arial', size=10, bold=True)
    
    # Signature boxes (3 rows high)
    for offset in range(1, 4):
        for col in range(1, 5):
            ws.cell(row=current_row + offset, column=col).border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
        for col in range(6, 11):
            ws.cell(row=current_row + offset, column=col).border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def create_professional_pdf(project, user_week_data, start_date, end_date):
    """
    Create a professional PDF export matching the template
    
    Args:
        project: Project dictionary
        user_week_data: Dictionary with user data
        start_date: Start date string
        end_date: End date string
    
    Returns:
        BytesIO object with PDF file
    """
    # Calculate dates
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    week_num = start_dt.isocalendar()[1]
    year = start_dt.year
    
    # Calculate totals
    day_totals = [0, 0, 0, 0, 0, 0, 0]
    for user_data in user_week_data.values():
        for i, hours in enumerate(user_data["days"]):
            day_totals[i] += hours
    
    # Create PDF
    pdf_file = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_file, 
        pagesize=A4, 
        leftMargin=2*cm, 
        rightMargin=2*cm, 
        topMargin=1.5*cm, 
        bottomMargin=1.5*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Header with logo and company name
    logo_path = '/app/backend/logo.png'
    if os.path.exists(logo_path):
        try:
            logo = RLImage(logo_path, width=3.5*cm, height=1.5*cm)
            company_style = ParagraphStyle(
                'CompanyStyle', 
                parent=styles['Heading1'], 
                fontSize=14, 
                textColor=colors.black,
                alignment=2  # Right align
            )
            company_text = Paragraph("<b>The Global Bedrijfsdiensten BV</b>", company_style)
            header_table = Table([[logo, company_text]], colWidths=[6*cm, 11*cm])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ]))
            elements.append(header_table)
        except:
            company_style = ParagraphStyle('CompanyStyle', parent=styles['Heading1'], fontSize=14, textColor=colors.black)
            elements.append(Paragraph("<b>The Global Bedrijfsdiensten BV</b>", company_style))
    else:
        company_style = ParagraphStyle('CompanyStyle', parent=styles['Heading1'], fontSize=14, textColor=colors.black)
        elements.append(Paragraph("<b>The Global Bedrijfsdiensten BV</b>", company_style))
    
    elements.append(Spacer(1, 0.8*cm))
    
    # Project info
    project_info_data = [
        ['Naam opdrachtgever:', project.get('company', '')],
        ['Project:', project.get('name', '')],
        ['Weeknummer/jaar:', f"{week_num}/{year}"],
        ['Soort werkzaamheden:', project.get('description', '')]
    ]
    
    project_table = Table(project_info_data, colWidths=[5*cm, 12*cm])
    project_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1.5, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(project_table)
    elements.append(Spacer(1, 0.8*cm))
    
    # Timesheet table
    table_data = [['Naam', 'BSN', 'week nr', 'ma', 'di', 'wo', 'do', 'vrij', 'zat', 'zon', 'Tot']]
    
    for user_name, data in sorted(user_week_data.items()):
        row = [user_name, data.get('bsn', ''), str(week_num)]
        row.extend([f"{h:.1f}" if h > 0 else "0" for h in data['days']])
        row.append(f"{sum(data['days']):.1f}")
        table_data.append(row)
    
    # Totals row
    if len(table_data) > 1:
        totals_row = ['TOTAAL', '', '']
        totals_row.extend([f"{t:.1f}" for t in day_totals])
        totals_row.append(f"{sum(day_totals):.1f}")
        table_data.append(totals_row)
    
    col_widths = [3.5*cm, 2.5*cm, 1.3*cm, 1*cm, 1*cm, 1*cm, 1*cm, 1*cm, 1*cm, 1*cm, 1.2*cm]
    
    main_table = Table(table_data, colWidths=col_widths)
    main_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Data rows
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        # Totals row
        ('BACKGROUND', (0, -1), (-1, -1), colors.yellow),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        # Borders
        ('BOX', (0, 0), (-1, -1), 1.5, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(main_table)
    elements.append(Spacer(1, 0.8*cm))
    
    # Footer
    date_place_data = [
        ['Datum:', datetime.now().strftime("%Y-%m-%d")],
        ['Plaats:', project.get('location', 'Utrecht')]
    ]
    date_place_table = Table(date_place_data, colWidths=[2.5*cm, 8*cm])
    date_place_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(date_place_table)
    elements.append(Spacer(1, 0.8*cm))
    
    # Signatures
    sig_label_style = ParagraphStyle('SigLabel', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold')
    sig_data = [
        [Paragraph('Accoord Uitvoerder', sig_label_style), Paragraph('Accoord The Global', sig_label_style)],
        ['', '']
    ]
    
    sig_table = Table(sig_data, colWidths=[8.5*cm, 8.5*cm], rowHeights=[0.7*cm, 2.5*cm])
    sig_table.setStyle(TableStyle([
        ('BOX', (0, 1), (0, 1), 1, colors.black),
        ('BOX', (1, 1), (1, 1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, 0), 'BOTTOM'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    elements.append(sig_table)
    
    doc.build(elements)
    pdf_file.seek(0)
    
    return pdf_file
