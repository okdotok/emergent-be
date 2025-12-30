from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi.responses import StreamingResponse
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
import calendar
from email_service import send_invitation_email, send_password_reset_email

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-change-this-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days
PASSWORD_RESET_EXPIRE_MINUTES = 60  # 1 hour

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    name: str
    role: str  # 'admin' or 'employee'
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    email: EmailStr
    name: str
    password: str
    invitation_token: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str
    user: User

class PasswordChange(BaseModel):
    old_password: str
    new_password: str

class PasswordResetRequest(BaseModel):
    email: EmailStr

class PasswordReset(BaseModel):
    token: str
    new_password: str

class PasswordResetToken(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    used: bool = False
    expires_at: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Invitation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    used: bool = False
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InvitationCreate(BaseModel):
    email: EmailStr

class Project(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProjectCreate(BaseModel):
    name: str
    description: Optional[str] = None

class TimeEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    project_id: str
    project_name: str
    date: str  # YYYY-MM-DD format
    hours: float
    note: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TimeEntryCreate(BaseModel):
    project_id: str
    date: str
    hours: float
    note: Optional[str] = None

# Helper functions
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except jwt.JWTError:
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user is None:
        raise HTTPException(status_code=401, detail="User not found")
    return User(**user)

async def get_admin_user(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

# Auth endpoints
@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserCreate):
    # Verify invitation
    invitation = await db.invitations.find_one(
        {"token": user_data.invitation_token, "used": False},
        {"_id": 0}
    )
    if not invitation or invitation["email"] != user_data.email:
        raise HTTPException(status_code=400, detail="Invalid or used invitation token")
    
    # Check if user already exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create user
    user = User(email=user_data.email, name=user_data.name, role="employee")
    user_dict = user.model_dump()
    user_dict["password"] = get_password_hash(user_data.password)
    user_dict["created_at"] = user_dict["created_at"].isoformat()
    
    await db.users.insert_one(user_dict)
    
    # Mark invitation as used
    await db.invitations.update_one(
        {"token": user_data.invitation_token},
        {"$set": {"used": True}}
    )
    
    # Create token
    access_token = create_access_token(data={"sub": user.id})
    
    return TokenResponse(access_token=access_token, token_type="bearer", user=user)

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Convert datetime
    if isinstance(user['created_at'], str):
        user['created_at'] = datetime.fromisoformat(user['created_at'])
    
    user_obj = User(**{k: v for k, v in user.items() if k != "password"})
    access_token = create_access_token(data={"sub": user_obj.id})
    
    return TokenResponse(access_token=access_token, token_type="bearer", user=user_obj)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

@api_router.post("/auth/change-password")
async def change_password(password_data: PasswordChange, current_user: User = Depends(get_current_user)):
    # Get user with password
    user = await db.users.find_one({"id": current_user.id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify old password
    if not verify_password(password_data.old_password, user["password"]):
        raise HTTPException(status_code=400, detail="Incorrect current password")
    
    # Update password
    new_password_hash = get_password_hash(password_data.new_password)
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"password": new_password_hash}}
    )
    
    return {"message": "Password changed successfully"}

@api_router.post("/auth/forgot-password")
async def forgot_password(request: PasswordResetRequest):
    # Check if user exists
    user = await db.users.find_one({"email": request.email}, {"_id": 0})
    if not user:
        # Don't reveal if user exists or not for security
        return {"message": "If the email exists, a password reset link has been sent"}
    
    # Create reset token
    reset_token = PasswordResetToken(
        user_id=user["id"],
        expires_at=datetime.now(timezone.utc) + timedelta(minutes=PASSWORD_RESET_EXPIRE_MINUTES)
    )
    reset_dict = reset_token.model_dump()
    reset_dict["created_at"] = reset_dict["created_at"].isoformat()
    reset_dict["expires_at"] = reset_dict["expires_at"].isoformat()
    
    await db.password_resets.insert_one(reset_dict)
    
    # Send email
    send_password_reset_email(request.email, reset_token.token)
    
    return {"message": "If the email exists, a password reset link has been sent"}

@api_router.post("/auth/reset-password")
async def reset_password(reset_data: PasswordReset):
    # Find reset token
    reset_token = await db.password_resets.find_one(
        {"token": reset_data.token, "used": False},
        {"_id": 0}
    )
    
    if not reset_token:
        raise HTTPException(status_code=400, detail="Invalid or used reset token")
    
    # Check if expired
    expires_at = datetime.fromisoformat(reset_token["expires_at"])
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="Reset token has expired")
    
    # Update password
    new_password_hash = get_password_hash(reset_data.new_password)
    await db.users.update_one(
        {"id": reset_token["user_id"]},
        {"$set": {"password": new_password_hash}}
    )
    
    # Mark token as used
    await db.password_resets.update_one(
        {"token": reset_data.token},
        {"$set": {"used": True}}
    )
    
    return {"message": "Password reset successfully"}

# Invitation endpoints
@api_router.post("/invitations", response_model=Invitation)
async def create_invitation(invitation_data: InvitationCreate, admin: User = Depends(get_admin_user)):
    # Check if email already invited and not used
    existing = await db.invitations.find_one(
        {"email": invitation_data.email, "used": False}
    )
    if existing:
        raise HTTPException(status_code=400, detail="Active invitation already exists for this email")
    
    invitation = Invitation(
        email=invitation_data.email,
        created_by=admin.id
    )
    invitation_dict = invitation.model_dump()
    invitation_dict["created_at"] = invitation_dict["created_at"].isoformat()
    
    await db.invitations.insert_one(invitation_dict)
    
    # Send email
    send_invitation_email(invitation_data.email, invitation.token)
    
    return invitation

@api_router.get("/invitations", response_model=List[Invitation])
async def get_invitations(admin: User = Depends(get_admin_user)):
    invitations = await db.invitations.find({}, {"_id": 0}).to_list(1000)
    for inv in invitations:
        if isinstance(inv['created_at'], str):
            inv['created_at'] = datetime.fromisoformat(inv['created_at'])
    return invitations

@api_router.get("/invitations/verify/{token}")
async def verify_invitation(token: str):
    invitation = await db.invitations.find_one(
        {"token": token, "used": False},
        {"_id": 0}
    )
    if not invitation:
        raise HTTPException(status_code=404, detail="Invalid or used invitation")
    return {"valid": True, "email": invitation["email"]}

# Project endpoints
@api_router.post("/projects", response_model=Project)
async def create_project(project_data: ProjectCreate, admin: User = Depends(get_admin_user)):
    project = Project(**project_data.model_dump())
    project_dict = project.model_dump()
    project_dict["created_at"] = project_dict["created_at"].isoformat()
    
    await db.projects.insert_one(project_dict)
    return project

@api_router.get("/projects", response_model=List[Project])
async def get_projects(current_user: User = Depends(get_current_user)):
    projects = await db.projects.find({"active": True}, {"_id": 0}).to_list(1000)
    for proj in projects:
        if isinstance(proj['created_at'], str):
            proj['created_at'] = datetime.fromisoformat(proj['created_at'])
    return projects

@api_router.put("/projects/{project_id}", response_model=Project)
async def update_project(project_id: str, project_data: ProjectCreate, admin: User = Depends(get_admin_user)):
    result = await db.projects.update_one(
        {"id": project_id},
        {"$set": project_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if isinstance(project['created_at'], str):
        project['created_at'] = datetime.fromisoformat(project['created_at'])
    return Project(**project)

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, admin: User = Depends(get_admin_user)):
    result = await db.projects.update_one(
        {"id": project_id},
        {"$set": {"active": False}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"success": True}

# Time entries endpoints
@api_router.post("/time-entries", response_model=TimeEntry)
async def create_time_entry(entry_data: TimeEntryCreate, current_user: User = Depends(get_current_user)):
    # Get project details
    project = await db.projects.find_one({"id": entry_data.project_id, "active": True}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    time_entry = TimeEntry(
        user_id=current_user.id,
        user_name=current_user.name,
        project_id=entry_data.project_id,
        project_name=project["name"],
        date=entry_data.date,
        hours=entry_data.hours,
        note=entry_data.note
    )
    entry_dict = time_entry.model_dump()
    entry_dict["created_at"] = entry_dict["created_at"].isoformat()
    
    await db.time_entries.insert_one(entry_dict)
    return time_entry

@api_router.get("/time-entries", response_model=List[TimeEntry])
async def get_time_entries(
    user_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    
    # Employees can only see their own entries
    if current_user.role == "employee":
        query["user_id"] = current_user.id
    elif user_id:  # Admin can filter by user
        query["user_id"] = user_id
    
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    entries = await db.time_entries.find(query, {"_id": 0}).to_list(10000)
    for entry in entries:
        if isinstance(entry['created_at'], str):
            entry['created_at'] = datetime.fromisoformat(entry['created_at'])
    return entries

@api_router.put("/time-entries/{entry_id}", response_model=TimeEntry)
async def update_time_entry(
    entry_id: str,
    entry_data: TimeEntryCreate,
    current_user: User = Depends(get_current_user)
):
    # Get existing entry
    entry = await db.time_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Time entry not found")
    
    # Check permissions
    if current_user.role == "employee" and entry["user_id"] != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get project details
    project = await db.projects.find_one({"id": entry_data.project_id, "active": True}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    update_data = {
        "project_id": entry_data.project_id,
        "project_name": project["name"],
        "date": entry_data.date,
        "hours": entry_data.hours,
        "note": entry_data.note
    }
    
    await db.time_entries.update_one({"id": entry_id}, {"$set": update_data})
    
    updated_entry = await db.time_entries.find_one({"id": entry_id}, {"_id": 0})
    if isinstance(updated_entry['created_at'], str):
        updated_entry['created_at'] = datetime.fromisoformat(updated_entry['created_at'])
    return TimeEntry(**updated_entry)

@api_router.delete("/time-entries/{entry_id}")
async def delete_time_entry(entry_id: str, current_user: User = Depends(get_current_user)):
    entry = await db.time_entries.find_one({"id": entry_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Time entry not found")
    
    if current_user.role == "employee" and entry["user_id"] != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.time_entries.delete_one({"id": entry_id})
    return {"success": True}

# Reports endpoints
@api_router.get("/reports/excel")
async def export_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_id: Optional[str] = None,
    admin: User = Depends(get_admin_user)
):
    # Build query
    query = {}
    if user_id:
        query["user_id"] = user_id
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    # Get entries
    entries = await db.time_entries.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Urenregistratie"
    
    # Style
    header_fill = PatternFill(start_color="16a085", end_color="16a085", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ["Datum", "Medewerker", "Project", "Uren", "Opmerking"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row_idx, entry in enumerate(entries, 2):
        ws.cell(row=row_idx, column=1, value=entry["date"])
        ws.cell(row=row_idx, column=2, value=entry["user_name"])
        ws.cell(row=row_idx, column=3, value=entry["project_name"])
        ws.cell(row=row_idx, column=4, value=entry["hours"])
        ws.cell(row=row_idx, column=5, value=entry.get("note", ""))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"urenregistratie_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/pdf-month")
async def export_pdf_month(
    year: int,
    month: int,
    user_id: Optional[str] = None,
    admin: User = Depends(get_admin_user)
):
    # Build query for the specific month
    start_date = f"{year}-{month:02d}-01"
    last_day = calendar.monthrange(year, month)[1]
    end_date = f"{year}-{month:02d}-{last_day}"
    
    query = {
        "date": {"$gte": start_date, "$lte": end_date}
    }
    if user_id:
        query["user_id"] = user_id
    
    # Get entries
    entries = await db.time_entries.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    
    # Create PDF
    pdf_file = io.BytesIO()
    doc = SimpleDocTemplate(pdf_file, pagesize=A4)
    
    # Container for elements
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    month_names = ['Januari', 'Februari', 'Maart', 'April', 'Mei', 'Juni',
                   'Juli', 'Augustus', 'September', 'Oktober', 'November', 'December']
    title = Paragraph(f"<b>Maandregister Uren - {month_names[month-1]} {year}</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    if not entries:
        no_data = Paragraph("Geen urenregistraties gevonden voor deze periode.", styles['Normal'])
        elements.append(no_data)
    else:
        # Create table data
        table_data = [['Datum', 'Medewerker', 'Project', 'Uren', 'Opmerking']]
        
        total_hours = 0
        for entry in entries:
            table_data.append([
                entry['date'],
                entry['user_name'],
                entry['project_name'],
                str(entry['hours']),
                entry.get('note', '')[:30]  # Truncate long notes
            ])
            total_hours += entry['hours']
        
        # Add total row
        table_data.append(['', '', 'Totaal', str(total_hours), ''])
        
        # Create table
        table = Table(table_data, colWidths=[3*cm, 4*cm, 5*cm, 2*cm, 5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16a085')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
    
    # Build PDF
    doc.build(elements)
    pdf_file.seek(0)
    
    filename = f"maandregister_{year}_{month:02d}.pdf"
    
    return StreamingResponse(
        pdf_file,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Users endpoint for admin
@api_router.get("/users", response_model=List[User])
async def get_users(admin: User = Depends(get_admin_user)):
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    for user in users:
        if isinstance(user['created_at'], str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
    return users

# Initialize first admin user
@api_router.post("/init-admin")
async def init_admin():
    # Check if any admin exists
    admin_exists = await db.users.find_one({"role": "admin"})
    if admin_exists:
        raise HTTPException(status_code=400, detail="Admin already exists")
    
    # Create default admin
    admin = User(
        email="admin@theglobal.nl",
        name="Administrator",
        role="admin"
    )
    admin_dict = admin.model_dump()
    admin_dict["password"] = get_password_hash("admin123")
    admin_dict["created_at"] = admin_dict["created_at"].isoformat()
    
    await db.users.insert_one(admin_dict)
    return {"message": "Admin created", "email": "admin@theglobal.nl", "password": "admin123"}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()