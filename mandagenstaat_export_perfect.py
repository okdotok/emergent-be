"""
PIXEL-PERFECT Mandagenstaat Export
100% identiek aan gebruiker template
"""

import io
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm


def create_perfect_excel(project, user_week_data, start_date, end_date):
    """
    Create PIXEL-PERFECT Excel matching template exactly
    - Exacte kolom breedtes
    - Exacte borders (medium op headers/totalen)
    - Arial 10pt
    - Datum format: dd-mm-yyyy
    - Logo groot en strak uitgelijnd
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "opl"
    
    # === EXACTE KOLOM BREEDTES (van template) ===
    ws.column_dimensions['A'].width = 5.453125
    ws.column_dimensions['B'].width = 21.453125
    ws.column_dimensions['C'].width = 11.542969
    ws.column_dimensions['D'].width = 12.726562
    ws.column_dimensions['E'].width = 4.453125
    ws.column_dimensions['F'].width = 8.43
    ws.column_dimensions['G'].width = 8.43
    ws.column_dimensions['H'].width = 8.43
    ws.column_dimensions['I'].width = 8.43
    ws.column_dimensions['J'].width = 8.43
    ws.column_dimensions['K'].width = 8.43
    ws.column_dimensions['L'].width = 5.0
    
    # === EXACTE RIJ HOOGTES ===
    ws.row_dimensions[10].height = 13.0
    ws.row_dimensions[14].height = 13.0
    ws.row_dimensions[17].height = 13.0
    ws.row_dimensions[18].height = 13.0
    ws.row_dimensions[34].height = 13.0
    ws.row_dimensions[35].height = 13.0
    
    # === LOGO (GROOT en strak uitgelijnd, zoals user wil) ===
    logo_path = '/app/backend/logo.png'
    if os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            # GROTER logo volgens user eis (ca. 2.5 kolommen breed, 9 rijen hoog)
            img.width = 250
            img.height = 234
            ws.add_image(img, 'A1')
        except Exception as e:
            print(f"Logo error: {e}")
    
    # === D9: Company Name (gecentreerd, Arial 10) ===
    ws['D9'] = "The Global Bedrijfsdiensten BV"
    ws['D9'].font = Font(name='Arial', size=10, bold=False)
    ws['D9'].alignment = Alignment(horizontal='center')
    # NO FILL - transparent/white background
    
    # === RIJEN 11-14: Project Info ===
    # Datum berekenen
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    week_num = start_dt.isocalendar()[1]
    year = start_dt.year
    
    info_rows = [
        (11, "Naam opdrachtgever", project.get('company', '')),
        (12, "Project", project.get('name', '')),
        (13, "Weeknummer/ jaar", f"{week_num}/{year}"),
        (14, "Soort werkzaamheden", project.get('description', ''))
    ]
    
    # EXACTE lichtgrijze borders zoals voorbeeld: #D9D9D9
    thin_border = Side(style='thin', color='D9D9D9')
    medium_border = Side(style='medium', color='D9D9D9')
    
    # Verschillende border styles voor verschillende secties
    # TABEL 1: Project info (rij 11-14) heeft DIKKE (medium) buitenranden
    # TABEL 2: Uren (rij 19-35) heeft DIKKE (medium) buitenranden
    
    # TABEL 1: Project info (rijen 11-14) met DIKKE buitenranden
    for idx, (row_num, label, value) in enumerate(info_rows):
        # B kolom: labels (dikgedrukt)
        ws.cell(row=row_num, column=2, value=label)
        ws.cell(row=row_num, column=2).font = Font(name='Arial', size=10, bold=True)
        
        # Borders voor B kolom (label): MEDIUM links/rechts
        if idx == 0:  # Eerste rij (row 11): MEDIUM top
            ws.cell(row=row_num, column=2).border = Border(
                left=medium_border, right=medium_border, top=medium_border, bottom=thin_border)
        elif idx == len(info_rows) - 1:  # Laatste rij (row 14): MEDIUM bottom
            ws.cell(row=row_num, column=2).border = Border(
                left=medium_border, right=medium_border, top=thin_border, bottom=medium_border)
        else:  # Middelste rijen: thin top/bottom
            ws.cell(row=row_num, column=2).border = Border(
                left=medium_border, right=medium_border, top=thin_border, bottom=thin_border)
        
        # C-L kolommen: waarde velden met MEDIUM rechter rand alleen
        for col in range(3, 13):  # C tot L
            cell = ws.cell(row=row_num, column=col)
            if col == 3:  # C kolom heeft de waarde
                cell.value = value
            cell.font = Font(name='Arial', size=10, bold=False)
            
            # Borders: thin links, MEDIUM rechts bij kolom L, thin/medium boven/onder
            if idx == 0:  # Eerste rij: MEDIUM top
                if col == 12:  # Kolom L: MEDIUM rechts
                    cell.border = Border(left=thin_border, right=medium_border, top=medium_border, bottom=thin_border)
                else:
                    cell.border = Border(left=thin_border, right=thin_border, top=medium_border, bottom=thin_border)
            elif idx == len(info_rows) - 1:  # Laatste rij: MEDIUM bottom
                if col == 12:  # Kolom L: MEDIUM rechts
                    cell.border = Border(left=thin_border, right=medium_border, top=thin_border, bottom=medium_border)
                else:
                    cell.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=medium_border)
            else:  # Middelste rijen
                if col == 12:  # Kolom L: MEDIUM rechts
                    cell.border = Border(left=thin_border, right=medium_border, top=thin_border, bottom=thin_border)
                else:
                    cell.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    # === RIJ 19: Headers TABEL 2 (uren) met DIKKE buitenranden ===
    headers = [
        (2, 'Naam ', None),
        (3, 'BSN', 'center'),
        (4, 'week nummer', 'center'),
        (5, 'ma ', None),
        (6, 'di', None),
        (7, 'wo', None),
        (8, 'do', None),
        (9, 'vrij ', None),
        (10, 'zat', None),
        (11, 'zon', None)
    ]
    
    for col, header_text, align in headers:
        cell = ws.cell(row=19, column=col, value=header_text)
        cell.font = Font(name='Arial', size=10, bold=True)
        
        # DIKKE (medium) buitenranden voor header row
        if col == 2:  # Eerste kolom: MEDIUM links
            cell.border = Border(left=medium_border, right=thin_border, top=medium_border, bottom=thin_border)
        elif col == 11:  # Laatste kolom: MEDIUM rechts
            cell.border = Border(left=thin_border, right=medium_border, top=medium_border, bottom=thin_border)
        else:  # Middelste kolommen: thin links/rechts, MEDIUM boven
            cell.border = Border(left=thin_border, right=thin_border, top=medium_border, bottom=thin_border)
        
        # NO FILL - transparant/wit zoals user template
        if align:
            cell.alignment = Alignment(horizontal=align)
    
    # === RIJEN 20-34: Data (15 rijen voor werknemers) met MEDIUM zijkanten ===
    current_row = 20
    first_data_row = True
    for user_name, data in sorted(user_week_data.items()):
        if current_row > 34:
            break
        
        # Eerste rij (20) heeft GEEN top border volgens template!
        top_style = None if first_data_row else thin_border
        
        # B: Naam (MEDIUM links)
        ws.cell(row=current_row, column=2, value=user_name)
        ws.cell(row=current_row, column=2).font = Font(name='Arial', size=10)
        ws.cell(row=current_row, column=2).border = Border(
            left=medium_border, right=thin_border, top=top_style, bottom=thin_border)
        
        # C: BSN
        ws.cell(row=current_row, column=3, value=data.get('bsn', ''))
        ws.cell(row=current_row, column=3).font = Font(name='Arial', size=10)
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=3).border = Border(
            left=thin_border, right=thin_border, top=top_style, bottom=thin_border)
        
        # D: Week nummer
        ws.cell(row=current_row, column=4, value=week_num)
        ws.cell(row=current_row, column=4).font = Font(name='Arial', size=10)
        ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=4).border = Border(
            left=thin_border, right=thin_border, top=top_style, bottom=thin_border)
        
        # E-J: ma-zat (dagen)
        for col_idx, hours in enumerate(data['days'][:6], start=5):
            cell = ws.cell(row=current_row, column=col_idx, value=hours if hours > 0 else 0)
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0.0'
            cell.border = Border(
                left=thin_border, right=thin_border, top=top_style, bottom=thin_border)
        
        # K: zondag (MEDIUM rechts)
        if len(data['days']) >= 7:
            cell = ws.cell(row=current_row, column=11, value=data['days'][6] if data['days'][6] > 0 else 0)
        else:
            cell = ws.cell(row=current_row, column=11, value=0)
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = '0.0'
        cell.border = Border(
            left=thin_border, right=medium_border, top=top_style, bottom=thin_border)
        
        first_data_row = False
        current_row += 1
    
    # === Vul lege rijen tot 34 met borders (voor volledige belijning) ===
    while current_row <= 34:
        # B: MEDIUM links
        ws.cell(row=current_row, column=2).border = Border(
            left=medium_border, right=thin_border, top=thin_border, bottom=thin_border)
        # C-J: thin rondom
        for col in range(3, 11):
            ws.cell(row=current_row, column=col).border = Border(
                left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        # K: MEDIUM rechts
        ws.cell(row=current_row, column=11).border = Border(
            left=thin_border, right=medium_border, top=thin_border, bottom=thin_border)
        current_row += 1
    
    # === RIJ 35: Totalen met DIKKE (medium) onderrand ===
    # B35: Empty met MEDIUM links/onder
    ws.cell(row=35, column=2).border = Border(
        left=medium_border, right=thin_border, top=thin_border, bottom=medium_border)
    
    # C35-D35: Empty met thin borders, MEDIUM onder
    for col in range(3, 5):  # C, D
        ws.cell(row=35, column=col).border = Border(
            left=thin_border, right=thin_border, top=thin_border, bottom=medium_border)
    
    # E35-J35: SUM formules (ma-zat)
    for col_idx, col_letter in enumerate(['E', 'F', 'G', 'H', 'I', 'J'], start=5):
        cell = ws.cell(row=35, column=col_idx)
        cell.value = f"=SUM({col_letter}20:{col_letter}34)"
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.border = Border(
            left=thin_border, right=thin_border, top=thin_border, bottom=medium_border)
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = '0.0'
    
    # K35: SUM zondag (thin rechts, MEDIUM onder - EXACT zoals template)
    ws.cell(row=35, column=11).value = "=SUM(K20:K34)"
    ws.cell(row=35, column=11).font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=35, column=11).border = Border(
        left=thin_border, right=thin_border, top=thin_border, bottom=medium_border)
    ws.cell(row=35, column=11).alignment = Alignment(horizontal='center')
    ws.cell(row=35, column=11).number_format = '0.0'
    
    # L35: Grand total (MEDIUM rondom - eigen vakje)
    ws.cell(row=35, column=12).value = "=SUM(E35:K35)"
    ws.cell(row=35, column=12).font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=35, column=12).border = Border(
        left=medium_border, right=medium_border, top=medium_border, bottom=medium_border)
    ws.cell(row=35, column=12).alignment = Alignment(horizontal='center')
    ws.cell(row=35, column=12).number_format = '0.0'
    
    # === RIJ 37-38: Datum en Plaats ===
    # DATUM = Runtime moment van uitdraaien (dd-mm-yyyy)
    ws['B37'] = "Datum: "
    ws['B37'].font = Font(name='Arial', size=10, bold=True)
    ws['B37'].alignment = Alignment(horizontal='left')
    # NO FILL
    
    # Runtime datum van dit moment (niet uit data!)
    ws['C37'] = datetime.now().strftime("%d-%m-%Y")  # Runtime: dd-mm-yyyy
    ws['C37'].font = Font(name='Arial', size=10)
    ws['C37'].alignment = Alignment(horizontal='left')
    # NO FILL
    
    # PLAATS = Locatie van uitdraai (niet werkplaats/project locatie)
    ws['B38'] = "Plaats: "
    ws['B38'].font = Font(name='Arial', size=10, bold=True)
    # NO FILL
    
    ws['C38'] = "Utrecht"  # Vaste plaats van uitdraai (niet project.location)
    ws['C38'].font = Font(name='Arial', size=10)
    # NO FILL
    
    # === RIJ 41: Accoord labels ===
    ws['B41'] = "Accoord Uitvoerder"
    ws['B41'].font = Font(name='Arial', size=10, bold=True)
    # NO FILL
    
    ws['E41'] = "Accoord The Global"
    ws['E41'].font = Font(name='Arial', size=10, bold=True)
    # NO FILL
    
    # === FREEZE PANES (header row blijft zichtbaar bij scrollen) ===
    ws.freeze_panes = 'B20'  # Freeze alles boven rij 20 (header blijft zichtbaar)
    
    # === PRINT SETTINGS ===
    # A4 formaat, normale marges, header repeat
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.print_title_rows = '19:19'  # Repeat header row
    ws.page_margins.left = 0.75
    ws.page_margins.right = 0.75
    ws.page_margins.top = 1.0
    ws.page_margins.bottom = 1.0
    ws.page_setup.scale = 100  # 100% scale
    
    # === Save to BytesIO ===
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def create_perfect_pdf(project, user_week_data, start_date, end_date):
    """
    Create PIXEL-PERFECT PDF matching template
    - Identieke layout als Excel
    - Logo groot en strak
    - Datum format: dd-mm-yyyy (runtime)
    - Medium borders op headers
    - Clean spacing
    - Embedded fonts (geen externe loads)
    """
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    week_num = start_dt.isocalendar()[1]
    year = start_dt.year
    
    # Calculate totals
    day_totals = [0, 0, 0, 0, 0, 0, 0]
    for user_data in user_week_data.values():
        for i, hours in enumerate(user_data["days"]):
            day_totals[i] += hours
    
    pdf_file = io.BytesIO()
    
    # SimpleDocTemplate met embed fonts
    doc = SimpleDocTemplate(
        pdf_file, 
        pagesize=A4, 
        leftMargin=1.5*cm, 
        rightMargin=1.5*cm, 
        topMargin=1.5*cm, 
        bottomMargin=1.5*cm,
        title="Mandagenstaat",
        author="The Global"
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # === LOGO + COMPANY NAME (GROOT en strak uitgelijnd zoals user wil) ===
    logo_path = '/app/backend/logo.png'
    if os.path.exists(logo_path):
        try:
            # GROTER logo volgens user eis
            logo = RLImage(logo_path, width=5*cm, height=3*cm)
            company_style = ParagraphStyle(
                'CompanyStyle', 
                parent=styles['Heading1'], 
                fontSize=12, 
                textColor=colors.black,
                alignment=2  # Right
            )
            company_text = Paragraph("<b>The Global Bedrijfsdiensten BV</b>", company_style)
            header_table = Table([[logo, company_text]], colWidths=[6*cm, 12*cm])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ]))
            elements.append(header_table)
        except:
            company_style = ParagraphStyle('CompanyStyle', parent=styles['Heading1'], fontSize=12)
            elements.append(Paragraph("<b>The Global Bedrijfsdiensten BV</b>", company_style))
    else:
        company_style = ParagraphStyle('CompanyStyle', parent=styles['Heading1'], fontSize=12)
        elements.append(Paragraph("<b>The Global Bedrijfsdiensten BV</b>", company_style))
    
    elements.append(Spacer(1, 0.8*cm))
    
    # === PROJECT INFO (met borders) ===
    project_info_data = [
        ['Naam opdrachtgever', project.get('company', '')],
        ['Project', project.get('name', '')],
        ['Weeknummer/ jaar', f"{week_num}/{year}"],
        ['Soort werkzaamheden', project.get('description', '')]
    ]
    
    project_table = Table(project_info_data, colWidths=[5*cm, 13*cm])
    project_table.setStyle(TableStyle([
        # DIKKE (1.5pt) buitenranden voor eerste tabel, thin (0.5pt) binnen
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#D9D9D9')),  # DIKKE buitenrand
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),  # Thin binnenranden
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),  # Labels bold
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(project_table)
    elements.append(Spacer(1, 0.8*cm))
    
    # === TIMESHEET TABLE ===
    table_data = [['Naam', 'BSN', 'week nr', 'ma', 'di', 'wo', 'do', 'vrij', 'zat', 'zon']]
    
    for user_name, data in sorted(user_week_data.items()):
        row = [user_name, data.get('bsn', ''), str(week_num)]
        row.extend([f"{h:.1f}" if h > 0 else "0.0" for h in data['days']])
        table_data.append(row)
    
    # Totals row
    if len(table_data) > 1:
        totals_row = ['TOTAAL', '', '']
        totals_row.extend([f"{t:.1f}" for t in day_totals])
        table_data.append(totals_row)
    
    col_widths = [4*cm, 2.5*cm, 1.5*cm, 1.2*cm, 1.2*cm, 1.2*cm, 1.2*cm, 1.2*cm, 1.2*cm, 1.2*cm]
    
    main_table = Table(table_data, colWidths=col_widths)
    # Gebruik alleen Helvetica (altijd embedded in PDF, geen externe fonts)
    main_table.setStyle(TableStyle([
        # Headers (NO BACKGROUND zoals user template)
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        # NO BACKGROUND - transparant/wit zoals user template
        # Data
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 10),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        # Totals (NO BACKGROUND zoals user template)
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        # NO BACKGROUND - transparant/wit zoals user template
        # Borders: DIKKE (1.5pt) buitenranden, thin (0.5pt) binnen
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#D9D9D9')),  # DIKKE buitenrand
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),  # Thin binnenranden
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(main_table)
    elements.append(Spacer(1, 0.8*cm))
    
    # === FOOTER ===
    # DATUM = Runtime moment van uitdraaien (dd-mm-yyyy)
    # PLAATS = Locatie van uitdraai (niet werkplaats)
    date_place_data = [
        ['Datum:', datetime.now().strftime("%d-%m-%Y")],  # Runtime: dd-mm-yyyy
        ['Plaats:', 'Utrecht']  # Vaste plaats van uitdraai
    ]
    date_place_table = Table(date_place_data, colWidths=[2*cm, 8*cm])
    date_place_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(date_place_table)
    elements.append(Spacer(1, 0.8*cm))
    
    # === SIGNATURES ===
    sig_label_style = ParagraphStyle('SigLabel', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold')
    sig_data = [
        [Paragraph('Accoord Uitvoerder', sig_label_style), Paragraph('Accoord The Global', sig_label_style)],
        ['', '']
    ]
    
    sig_table = Table(sig_data, colWidths=[8.5*cm, 8.5*cm], rowHeights=[0.7*cm, 2.5*cm])
    sig_table.setStyle(TableStyle([
        # Borders: lichtgrijs #D9D9D9 voor handtekening vakken
        ('BOX', (0, 1), (0, 1), 0.5, colors.HexColor('#D9D9D9')),
        ('BOX', (1, 1), (1, 1), 0.5, colors.HexColor('#D9D9D9')),
        ('VALIGN', (0, 0), (-1, 0), 'BOTTOM'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    elements.append(sig_table)
    
    doc.build(elements)
    pdf_file.seek(0)
    
    return pdf_file
