from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi.responses import StreamingResponse
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
import calendar
from mandagenstaat_template_based import create_from_template, create_pdf_from_template

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
# In production, MONGO_URL must be set - don't default to localhost
mongo_url = os.environ.get('MONGO_URL')
if not mongo_url:
    raise ValueError("MONGO_URL environment variable is required. Please set it in your deployment configuration.")
db_name = os.environ.get('DB_NAME', 'theglobal_uren')
client = AsyncIOMotorClient(mongo_url)
db = client[db_name]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-change-this-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days
PASSWORD_RESET_EXPIRE_MINUTES = 60  # 1 hour

# Settings
DEFAULT_PROJECT_MATCH_RADIUS = 250  # meters

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Startup event to create indexes and ensure dependencies
@app.on_event("startup")
async def startup_event():
    """Create database indexes and ensure system dependencies"""
    import shutil
    
    # 1. Check for ssconvert (gnumeric) - don't try to install on Render/cloud platforms
    # Installation requires sudo which is not available on most cloud platforms
    try:
        if shutil.which("ssconvert"):
            print("✅ ssconvert available for PDF generation")
        else:
            print("⚠️  ssconvert not found - PDF export features may be limited")
            print("   Note: Install gnumeric manually if needed (requires system admin)")
    except Exception as e:
        print(f"⚠️  Gnumeric check warning: {e}")
    
    # 2. Create database indexes for performance
    try:
        await db.clock_entries.create_index([("user_id", 1), ("clock_in_time", -1)])
        await db.clock_entries.create_index([("id", 1)])
        print("✅ Database indexes created successfully")
    except Exception as e:
        print(f"⚠️  Index creation warning: {e}")

# Models
class Location(BaseModel):
    latitude: float
    longitude: float
    accuracy: Optional[float] = None
    address: Optional[str] = None

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    first_name: str
    last_name: str
    role: str  # 'admin' or 'employee'
    bsn: Optional[str] = None  # Burgerservicenummer (Dutch national ID)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    
    @property
    def full_name(self) -> str:
        return f"{self.first_name} {self.last_name}"

class UserCreate(BaseModel):
    email: EmailStr
    first_name: str
    last_name: str
    password: str
    invitation_token: str

class UserUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[EmailStr] = None
    bsn: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str
    user: User

class PasswordChange(BaseModel):
    old_password: str
    new_password: str

class PasswordResetRequest(BaseModel):
    email: EmailStr

class PasswordReset(BaseModel):
    token: str
    new_password: str

class PasswordResetToken(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    used: bool = False
    expires_at: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Invitation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    name: Optional[str] = None  # Naam van uitgenodigde persoon
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    used: bool = False
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InvitationCreate(BaseModel):
    email: EmailStr
    name: Optional[str] = None

class Project(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    company: str  # Bedrijf
    location: str  # Locatie/adres
    latitude: Optional[float] = None  # GPS coordinaat
    longitude: Optional[float] = None  # GPS coordinaat
    location_radius: float = 100.0  # Toegestane afwijking in meters
    description: Optional[str] = None
    active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProjectCreate(BaseModel):
    name: str
    company: str
    location: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    location_radius: Optional[float] = 100.0
    description: Optional[str] = None

class ClockEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    project_id: str
    project_name: str
    company: str
    project_location: str
    
    # Clock in data
    clock_in_time: datetime
    clock_in_location: Location
    
    # Clock out data (optional until clocked out)
    clock_out_time: Optional[datetime] = None
    clock_out_location: Optional[Location] = None
    clock_out_distance_m: Optional[float] = None
    clock_out_match: Optional[bool] = None
    clock_out_warning: Optional[str] = None
    
    # Calculated
    total_hours: Optional[float] = None
    status: str = "clocked_in"  # clocked_in, clocked_out
    
    # Location verification (clock in)
    location_warning: Optional[str] = None
    distance_to_project_m: Optional[float] = None  # Distance in meters to project location
    project_match: Optional[bool] = None  # True if within radius, False otherwise
    
    note: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ClockInRequest(BaseModel):
    project_id: str
    location: Location
    note: Optional[str] = None

class ClockOutRequest(BaseModel):
    location: Location
    note: Optional[str] = None

class BulkDeleteRequest(BaseModel):
    invitation_ids: List[str]

class AuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    action: str  # "delete_invitation", "bulk_delete_invitations", etc.
    performed_by: str  # user_id
    performed_by_name: str  # user name
    target_type: str  # "invitation", "user", etc.
    target_ids: List[str]  # IDs of deleted items
    target_emails: Optional[List[str]] = None  # For invitations
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    details: Optional[str] = None

class GPSCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    entry_id: str  # clock entry ID
    user_id: str
    user_name: str
    project_id: str
    location: Location
    distance_to_project_m: float
    check_time: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    check_type: str  # "clock_in", "periodic", "clock_out"

# Helper functions
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def calculate_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """
    Calculate distance between two GPS coordinates using Haversine formula
    Returns distance in meters
    """
    from math import radians, sin, cos, sqrt, atan2
    
    R = 6371000  # Earth radius in meters
    
    lat1_rad = radians(lat1)
    lat2_rad = radians(lat2)
    delta_lat = radians(lat2 - lat1)
    delta_lon = radians(lon2 - lon1)
    
    a = sin(delta_lat / 2) ** 2 + cos(lat1_rad) * cos(lat2_rad) * sin(delta_lon / 2) ** 2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))
    
    distance = R * c
    return distance

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except jwt.JWTError:
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user is None:
        raise HTTPException(status_code=401, detail="User not found")
    return User(**user)

async def get_admin_user(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

# Auth endpoints
@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserCreate):
    invitation = await db.invitations.find_one(
        {"token": user_data.invitation_token, "used": False},
        {"_id": 0}
    )
    if not invitation or invitation["email"] != user_data.email:
        raise HTTPException(status_code=400, detail="Invalid or used invitation token")
    
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user = User(email=user_data.email, first_name=user_data.first_name, last_name=user_data.last_name, role="employee")
    user_dict = user.model_dump()
    user_dict["password"] = get_password_hash(user_data.password)
    user_dict["created_at"] = user_dict["created_at"].isoformat()
    
    await db.users.insert_one(user_dict)
    await db.invitations.update_one(
        {"token": user_data.invitation_token},
        {"$set": {"used": True}}
    )
    
    access_token = create_access_token(data={"sub": user.id})
    return TokenResponse(access_token=access_token, token_type="bearer", user=user)

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if isinstance(user['created_at'], str):
        user['created_at'] = datetime.fromisoformat(user['created_at'])
    
    user_obj = User(**{k: v for k, v in user.items() if k != "password"})
    access_token = create_access_token(data={"sub": user_obj.id})
    
    return TokenResponse(access_token=access_token, token_type="bearer", user=user_obj)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

@api_router.post("/auth/change-password")
async def change_password(password_data: PasswordChange, current_user: User = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user.id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if not verify_password(password_data.old_password, user["password"]):
        raise HTTPException(status_code=400, detail="Incorrect current password")
    
    new_password_hash = get_password_hash(password_data.new_password)
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"password": new_password_hash}}
    )
    
    # Audit log
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "action": "password_change",
        "user_id": current_user.id,
        "user_name": f"{current_user.first_name} {current_user.last_name}",
        "details": "User changed own password"
    })
    
    return {"message": "Password changed successfully"}

@api_router.post("/admin/users/{user_id}/reset-password")
async def admin_reset_user_password(
    user_id: str,
    new_password: str,
    admin: User = Depends(get_admin_user)
):
    """Admin can reset any user's password (with audit log)"""
    # Check if target user exists
    target_user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Hash new password
    new_password_hash = get_password_hash(new_password)
    
    # Update password
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"password": new_password_hash}}
    )
    
    # Audit log
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "action": "admin_password_reset",
        "admin_id": admin.id,
        "admin_name": f"{admin.first_name} {admin.last_name}",
        "target_user_id": user_id,
        "target_user_name": f"{target_user['first_name']} {target_user['last_name']}",
        "details": f"Admin {admin.first_name} {admin.last_name} reset password for {target_user['first_name']} {target_user['last_name']}"
    })
    
    return {"message": "Password reset successfully", "audit_logged": True}

@api_router.post("/auth/forgot-password")
async def forgot_password(request: PasswordResetRequest):
    user = await db.users.find_one({"email": request.email}, {"_id": 0})
    if not user:
        return {"message": "If the email exists, a password reset link has been sent"}
    
    reset_token = PasswordResetToken(
        user_id=user["id"],
        expires_at=datetime.now(timezone.utc) + timedelta(minutes=PASSWORD_RESET_EXPIRE_MINUTES)
    )
    reset_dict = reset_token.model_dump()
    reset_dict["created_at"] = reset_dict["created_at"].isoformat()
    reset_dict["expires_at"] = reset_dict["expires_at"].isoformat()
    
    await db.password_resets.insert_one(reset_dict)
    
    # Send password reset email
    from email_service import send_password_reset_email
    send_password_reset_email(request.email, reset_token.token)
    
    return {"message": "If the email exists, a password reset link has been sent"}

@api_router.post("/auth/reset-password")
async def reset_password(reset_data: PasswordReset):
    reset_token = await db.password_resets.find_one(
        {"token": reset_data.token, "used": False},
        {"_id": 0}
    )
    
    if not reset_token:
        raise HTTPException(status_code=400, detail="Invalid or used reset token")
    
    expires_at = datetime.fromisoformat(reset_token["expires_at"])
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="Reset token has expired")
    
    new_password_hash = get_password_hash(reset_data.new_password)
    await db.users.update_one(
        {"id": reset_token["user_id"]},
        {"$set": {"password": new_password_hash}}
    )
    
    await db.password_resets.update_one(
        {"token": reset_data.token},
        {"$set": {"used": True}}
    )
    
    return {"message": "Password reset successfully"}

# Invitation endpoints
@api_router.post("/invitations", response_model=Invitation)
async def create_invitation(invitation_data: InvitationCreate, admin: User = Depends(get_admin_user)):
    existing = await db.invitations.find_one(
        {"email": invitation_data.email, "used": False}
    )
    if existing:
        raise HTTPException(status_code=400, detail="Active invitation already exists for this email")
    
    invitation = Invitation(
        email=invitation_data.email,
        name=invitation_data.name,
        created_by=admin.id
    )
    invitation_dict = invitation.model_dump()
    invitation_dict["created_at"] = invitation_dict["created_at"].isoformat()
    
    await db.invitations.insert_one(invitation_dict)
    
    # Send email
    from email_service import send_invitation_email
    success = send_invitation_email(invitation_data.email, invitation.token)
    
    if not success:
        # Delete the invitation if email fails
        await db.invitations.delete_one({"id": invitation.id})
        raise HTTPException(
            status_code=500, 
            detail="Failed to send invitation email. Please check the email address or contact support@transip.nl if the issue persists."
        )
    
    return invitation

@api_router.put("/invitations/{invitation_id}")
async def update_invitation(invitation_id: str, update_data: InvitationCreate, admin: User = Depends(get_admin_user)):
    """Update invitation name"""
    invitation = await db.invitations.find_one({"id": invitation_id}, {"_id": 0})
    if not invitation:
        raise HTTPException(status_code=404, detail="Invitation not found")
    
    # Update only name field
    await db.invitations.update_one(
        {"id": invitation_id},
        {"$set": {"name": update_data.name}}
    )
    
    return {"message": "Invitation updated successfully"}

@api_router.post("/invitations/{invitation_id}/resend")
async def resend_invitation(invitation_id: str, admin: User = Depends(get_admin_user)):
    # Get invitation
    invitation = await db.invitations.find_one({"id": invitation_id, "used": False}, {"_id": 0})
    if not invitation:
        raise HTTPException(status_code=404, detail="Invitation not found or already used")
    
    # Resend email
    from email_service import send_invitation_email
    success = send_invitation_email(invitation["email"], invitation["token"])
    
    if success:
        return {"message": "Invitation resent successfully"}
    else:
        raise HTTPException(status_code=500, detail="Failed to send email")

@api_router.delete("/invitations/{invitation_id}")
async def delete_invitation(invitation_id: str, admin: User = Depends(get_admin_user)):
    # Get invitation before deleting for audit log
    invitation = await db.invitations.find_one({"id": invitation_id}, {"_id": 0})
    if not invitation:
        raise HTTPException(status_code=404, detail="Invitation not found")
    
    # Delete invitation
    result = await db.invitations.delete_one({"id": invitation_id})
    
    # Create audit log
    audit = AuditLog(
        action="delete_invitation",
        performed_by=admin.id,
        performed_by_name=admin.full_name,
        target_type="invitation",
        target_ids=[invitation_id],
        target_emails=[invitation["email"]],
        details=f"Deleted invitation for {invitation['email']}"
    )
    audit_dict = audit.model_dump()
    audit_dict["timestamp"] = audit_dict["timestamp"].isoformat()
    await db.audit_logs.insert_one(audit_dict)
    
    return {"message": "Invitation deleted successfully"}

@api_router.get("/invitations", response_model=List[Invitation])
async def get_invitations(admin: User = Depends(get_admin_user)):
    invitations = await db.invitations.find({}, {"_id": 0}).to_list(1000)
    for inv in invitations:
        if isinstance(inv['created_at'], str):
            inv['created_at'] = datetime.fromisoformat(inv['created_at'])
    return invitations

@api_router.post("/invitations/bulk-delete")
async def bulk_delete_invitations(request: BulkDeleteRequest, admin: User = Depends(get_admin_user)):
    if not request.invitation_ids:
        raise HTTPException(status_code=400, detail="No invitation IDs provided")
    
    # Get invitations before deleting for audit log
    invitations = await db.invitations.find(
        {"id": {"$in": request.invitation_ids}}, 
        {"_id": 0}
    ).to_list(1000)
    
    if not invitations:
        raise HTTPException(status_code=404, detail="No invitations found")
    
    # Delete invitations
    result = await db.invitations.delete_many({"id": {"$in": request.invitation_ids}})
    
    # Create audit log
    deleted_emails = [inv["email"] for inv in invitations]
    audit = AuditLog(
        action="bulk_delete_invitations",
        performed_by=admin.id,
        performed_by_name=admin.full_name,
        target_type="invitation",
        target_ids=request.invitation_ids,
        target_emails=deleted_emails,
        details=f"Bulk deleted {result.deleted_count} invitations: {', '.join(deleted_emails)}"
    )
    audit_dict = audit.model_dump()
    audit_dict["timestamp"] = audit_dict["timestamp"].isoformat()
    await db.audit_logs.insert_one(audit_dict)
    
    return {
        "message": f"Successfully deleted {result.deleted_count} invitations",
        "deleted_count": result.deleted_count
    }

@api_router.get("/invitations/verify/{token}")
async def verify_invitation(token: str):
    invitation = await db.invitations.find_one(
        {"token": token, "used": False},
        {"_id": 0}
    )
    if not invitation:
        raise HTTPException(status_code=404, detail="Invalid or used invitation")
    return {"valid": True, "email": invitation["email"]}

# Project endpoints
@api_router.post("/projects", response_model=Project)
async def create_project(project_data: ProjectCreate, admin: User = Depends(get_admin_user)):
    project = Project(**project_data.model_dump())
    project_dict = project.model_dump()
    project_dict["created_at"] = project_dict["created_at"].isoformat()
    
    await db.projects.insert_one(project_dict)
    return project

@api_router.get("/projects", response_model=List[Project])
async def get_projects(current_user: User = Depends(get_current_user)):
    projects = await db.projects.find({"active": True}, {"_id": 0}).to_list(1000)
    for proj in projects:
        if isinstance(proj['created_at'], str):
            proj['created_at'] = datetime.fromisoformat(proj['created_at'])
    return projects

@api_router.put("/projects/{project_id}", response_model=Project)
async def update_project(project_id: str, project_data: ProjectCreate, admin: User = Depends(get_admin_user)):
    result = await db.projects.update_one(
        {"id": project_id},
        {"$set": project_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if isinstance(project['created_at'], str):
        project['created_at'] = datetime.fromisoformat(project['created_at'])
    return Project(**project)

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, admin: User = Depends(get_admin_user)):
    result = await db.projects.update_one(
        {"id": project_id},
        {"$set": {"active": False}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"success": True}

# Clock entries endpoints
@api_router.post("/clock/in", response_model=ClockEntry)
async def clock_in(clock_data: ClockInRequest, current_user: User = Depends(get_current_user)):
    # Check if user is already clocked in
    active_entry = await db.clock_entries.find_one({
        "user_id": current_user.id,
        "status": "clocked_in"
    })
    if active_entry:
        raise HTTPException(status_code=400, detail="Je bent al ingeklokt. Klok eerst uit voordat je opnieuw inklokt.")
    
    # Get project details
    project = await db.projects.find_one({"id": clock_data.project_id, "active": True}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # GPS Verification - STRICT 50m check
    location_warning = None
    distance_to_project = None
    project_match = None
    
    if not project.get("latitude") or not project.get("longitude"):
        raise HTTPException(status_code=400, detail="Project heeft geen GPS locatie ingesteld")
    
    distance_to_project = calculate_distance(
        clock_data.location.latitude,
        clock_data.location.longitude,
        project["latitude"],
        project["longitude"]
    )
    
    # STRICT: Must be within 50 meters
    if distance_to_project > 50:
        raise HTTPException(
            status_code=403, 
            detail=f"Te ver van project locatie: {int(distance_to_project)}m (maximaal 50m toegestaan)"
        )
    
    # Check if within DEFAULT_PROJECT_MATCH_RADIUS (250m) for reporting
    project_match = distance_to_project <= DEFAULT_PROJECT_MATCH_RADIUS
    
    radius = project.get("location_radius", 100.0)
    if distance_to_project > radius:
        location_warning = f"WAARSCHUWING: Locatie afwijking {int(distance_to_project)}m"
    
    clock_entry = ClockEntry(
        user_id=current_user.id,
        user_name=current_user.full_name,
        project_id=clock_data.project_id,
        project_name=project["name"],
        company=project["company"],
        project_location=project["location"],
        clock_in_time=datetime.now(timezone.utc),
        clock_in_location=clock_data.location,
        note=clock_data.note,
        status="clocked_in",
        location_warning=location_warning,
        distance_to_project_m=distance_to_project,
        project_match=project_match
    )
    
    entry_dict = clock_entry.model_dump()
    entry_dict["created_at"] = entry_dict["created_at"].isoformat()
    entry_dict["clock_in_time"] = entry_dict["clock_in_time"].isoformat()
    entry_dict["clock_in_location"] = clock_data.location.model_dump()
    
    await db.clock_entries.insert_one(entry_dict)
    return clock_entry

@api_router.post("/clock/out/{entry_id}", response_model=ClockEntry)
async def clock_out(entry_id: str, clock_data: ClockOutRequest, current_user: User = Depends(get_current_user)):
    # Get entry
    entry = await db.clock_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Clock entry not found")
    
    if entry["user_id"] != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    if entry["status"] == "clocked_out":
        raise HTTPException(status_code=400, detail="Already clocked out")
    
    # Get project for GPS validation
    project = await db.projects.find_one({"id": entry["project_id"]}, {"_id": 0})
    
    # Calculate distance for clock-out location
    clock_out_distance = None
    clock_out_match = None
    clock_out_warning = None
    
    if project and project.get("location_lat") and project.get("location_lon") and clock_data.location:
        from math import radians, sin, cos, sqrt, atan2
        
        lat1, lon1 = radians(project["location_lat"]), radians(project["location_lon"])
        lat2, lon2 = radians(clock_data.location.lat), radians(clock_data.location.lon)
        
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
        c = 2 * atan2(sqrt(a), sqrt(1-a))
        distance = 6371000 * c  # Earth radius in meters
        
        clock_out_distance = distance
        radius = project.get("location_radius", 50)
        clock_out_match = distance <= radius
        
        if not clock_out_match:
            clock_out_warning = f"Uitklok locatie is {int(distance)}m van project (max {radius}m)"
    
    # Calculate hours
    clock_in_time = datetime.fromisoformat(entry["clock_in_time"])
    clock_out_time = datetime.now(timezone.utc)
    time_diff = clock_out_time - clock_in_time
    total_hours = round(time_diff.total_seconds() / 3600, 2)
    
    # Update entry
    update_data = {
        "clock_out_time": clock_out_time.isoformat(),
        "clock_out_location": clock_data.location.model_dump(),
        "clock_out_distance_m": clock_out_distance,
        "clock_out_match": clock_out_match,
        "clock_out_warning": clock_out_warning,
        "total_hours": total_hours,
        "status": "clocked_out"
    }
    if clock_data.note:
        update_data["note"] = clock_data.note
    
    await db.clock_entries.update_one(
        {"id": entry_id},
        {"$set": update_data}
    )
    
    updated_entry = await db.clock_entries.find_one({"id": entry_id}, {"_id": 0})
    if isinstance(updated_entry['created_at'], str):
        updated_entry['created_at'] = datetime.fromisoformat(updated_entry['created_at'])
    if isinstance(updated_entry['clock_in_time'], str):
        updated_entry['clock_in_time'] = datetime.fromisoformat(updated_entry['clock_in_time'])
    if isinstance(updated_entry['clock_out_time'], str):
        updated_entry['clock_out_time'] = datetime.fromisoformat(updated_entry['clock_out_time'])
    
    return ClockEntry(**updated_entry)

@api_router.post("/clock/gps-log/{entry_id}")
async def log_gps_position(entry_id: str, location: Location, current_user: User = Depends(get_current_user)):
    """Log GPS position during active clock session"""
    from math import radians, sin, cos, sqrt, atan2
    
    # Get entry
    entry = await db.clock_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry or entry["user_id"] != current_user.id:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    if entry["status"] != "clocked_in":
        raise HTTPException(status_code=400, detail="Entry not active")
    
    # Get project for distance calc
    project = await db.projects.find_one({"id": entry["project_id"]}, {"_id": 0})
    
    distance = None
    within_radius = None
    
    if project and project.get("location_lat") and project.get("location_lon"):
        lat1, lon1 = radians(project["location_lat"]), radians(project["location_lon"])
        lat2, lon2 = radians(location.lat), radians(location.lon)
        
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
        c = 2 * atan2(sqrt(a), sqrt(1-a))
        distance = 6371000 * c
        
        radius = project.get("location_radius", 50)
        within_radius = distance <= radius
    
    # Store GPS log
    gps_log = {
        "id": str(uuid.uuid4()),
        "entry_id": entry_id,
        "user_id": current_user.id,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "location": location.model_dump(),
        "distance_to_project": distance,
        "within_radius": within_radius
    }
    
    await db.gps_logs.insert_one(gps_log)
    
    return {"success": True, "distance": distance, "within_radius": within_radius}

@api_router.get("/clock/status")
async def get_clock_status(current_user: User = Depends(get_current_user)):
    # Check if user has active clock entry
    active_entry = await db.clock_entries.find_one({
        "user_id": current_user.id,
        "status": "clocked_in"
    }, {"_id": 0})
    
    if active_entry:
        if isinstance(active_entry['created_at'], str):
            active_entry['created_at'] = datetime.fromisoformat(active_entry['created_at'])
        if isinstance(active_entry['clock_in_time'], str):
            active_entry['clock_in_time'] = datetime.fromisoformat(active_entry['clock_in_time'])
        return {"clocked_in": True, "entry": ClockEntry(**active_entry)}
    
    return {"clocked_in": False, "entry": None}

@api_router.get("/clock/entries")
async def get_clock_entries(
    user_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    date: Optional[str] = None,  # NEW: Exact date filter (YYYY-MM-DD)
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Get clock entries with strict filtering
    
    Parameters:
    - date: Exact date (YYYY-MM-DD) - shows ONLY entries from this date
    - start_date + end_date: Date range (inclusive)
    - user_id: Filter by specific user
    - status: Filter by entry status
    
    Timezone: All dates are normalized to Europe/Amsterdam
    """
    from zoneinfo import ZoneInfo
    
    query = {}
    
    # Employees can only see their own entries
    if current_user.role == "employee":
        query["user_id"] = current_user.id
    elif user_id:
        query["user_id"] = user_id
    
    if status:
        query["status"] = status
    
    # STRICT DATE FILTERING
    if date:
        # Exact date match - ONLY this date
        query["$expr"] = {
            "$eq": [{"$substr": ["$clock_in_time", 0, 10]}, date]
        }
    elif start_date or end_date:
        # Date range (inclusive boundaries)
        date_conditions = []
        if start_date:
            date_conditions.append({"$gte": [{"$substr": ["$clock_in_time", 0, 10]}, start_date]})
        if end_date:
            date_conditions.append({"$lte": [{"$substr": ["$clock_in_time", 0, 10]}, end_date]})
        
        if date_conditions:
            query["$expr"] = {"$and": date_conditions} if len(date_conditions) > 1 else date_conditions[0]
    
    entries = await db.clock_entries.find(query, {"_id": 0}).sort("clock_in_time", -1).to_list(10000)
    
    result = []
    for entry in entries:
        if isinstance(entry['created_at'], str):
            entry['created_at'] = datetime.fromisoformat(entry['created_at'])
        if isinstance(entry['clock_in_time'], str):
            entry['clock_in_time'] = datetime.fromisoformat(entry['clock_in_time'])
        if entry.get('clock_out_time') and isinstance(entry['clock_out_time'], str):
            entry['clock_out_time'] = datetime.fromisoformat(entry['clock_out_time'])
        result.append(ClockEntry(**entry))
    
    return result

@api_router.get("/clock/entries/{entry_id}")
async def get_single_clock_entry(
    entry_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get a single clock entry by ID"""
    entry = await db.clock_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Clock entry not found")
    
    # Employees can only see their own entries
    if current_user.role == "employee" and entry["user_id"] != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Convert datetime strings to datetime objects
    if isinstance(entry['created_at'], str):
        entry['created_at'] = datetime.fromisoformat(entry['created_at'])
    if isinstance(entry['clock_in_time'], str):
        entry['clock_in_time'] = datetime.fromisoformat(entry['clock_in_time'])
    if entry.get('clock_out_time') and isinstance(entry['clock_out_time'], str):
        entry['clock_out_time'] = datetime.fromisoformat(entry['clock_out_time'])
    
    return ClockEntry(**entry)

@api_router.delete("/clock/entries/{entry_id}")
async def delete_clock_entry(entry_id: str, current_user: User = Depends(get_current_user)):
    entry = await db.clock_entries.find_one({"id": entry_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Clock entry not found")
    
    if current_user.role == "employee" and entry["user_id"] != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.clock_entries.delete_one({"id": entry_id})
    return {"success": True}

@api_router.get("/admin/time-entries/overview")
async def get_admin_overview(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    project_ids: Optional[str] = None,  # comma-separated IDs
    user_ids: Optional[str] = None,  # comma-separated IDs
    admin: User = Depends(get_admin_user)
):
    """
    Get admin overview with filters and totals
    """
    query = {}
    
    # Date filter
    if start_date or end_date:
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        if date_query:
            query["$expr"] = {
                "$and": [
                    {"$gte": [{"$substr": ["$clock_in_time", 0, 10]}, date_query.get("$gte", "1900-01-01")]},
                    {"$lte": [{"$substr": ["$clock_in_time", 0, 10]}, date_query.get("$lte", "9999-12-31")]}
                ]
            }
    
    # Project filter
    if project_ids:
        query["project_id"] = {"$in": project_ids.split(",")}
    
    # User filter
    if user_ids:
        query["user_id"] = {"$in": user_ids.split(",")}
    
    # Get all entries matching filter - SORT OLD TO NEW
    entries = await db.clock_entries.find(query, {"_id": 0}).sort("clock_in_time", 1).to_list(10000)
    
    # Calculate totals
    total_hours = 0
    hours_per_project = {}
    hours_per_user = {}
    
    processed_entries = []
    for entry in entries:
        # Parse dates
        if isinstance(entry.get('created_at'), str):
            entry['created_at'] = datetime.fromisoformat(entry['created_at'])
        if isinstance(entry.get('clock_in_time'), str):
            entry['clock_in_time'] = datetime.fromisoformat(entry['clock_in_time'])
        if entry.get('clock_out_time') and isinstance(entry['clock_out_time'], str):
            entry['clock_out_time'] = datetime.fromisoformat(entry['clock_out_time'])
        
        # Calculate hours
        hours = entry.get('total_hours', 0) or 0
        total_hours += hours
        
        # Hours per project
        project_name = entry.get('project_name', 'Unknown')
        if project_name not in hours_per_project:
            hours_per_project[project_name] = 0
        hours_per_project[project_name] += hours
        
        # Hours per user
        user_name = entry.get('user_name', 'Unknown')
        if user_name not in hours_per_user:
            hours_per_user[user_name] = 0
        hours_per_user[user_name] += hours
        
        processed_entries.append(entry)
    
    # Sort per_user by hours (top 10)
    top_users = sorted(hours_per_user.items(), key=lambda x: x[1], reverse=True)[:10]
    
    return {
        "entries": processed_entries,
        "total_hours": round(total_hours, 2),
        "hours_per_project": hours_per_project,
        "hours_per_user_top10": dict(top_users),
        "hours_per_user_all": hours_per_user,
        "entry_count": len(processed_entries)
    }

@api_router.get("/time-entries/my-overview")
async def get_my_overview(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    project_ids: Optional[str] = None,  # comma-separated IDs
    current_user: User = Depends(get_current_user)
):
    """
    Get employee's own overview with filters and totals
    """
    query = {"user_id": current_user.id}
    
    # Date filter
    if start_date or end_date:
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        if date_query:
            query["$expr"] = {
                "$and": [
                    {"$gte": [{"$substr": ["$clock_in_time", 0, 10]}, date_query.get("$gte", "1900-01-01")]},
                    {"$lte": [{"$substr": ["$clock_in_time", 0, 10]}, date_query.get("$lte", "9999-12-31")]}
                ]
            }
    
    # Project filter
    if project_ids:
        query["project_id"] = {"$in": project_ids.split(",")}
    
    # Get entries - SORT OLD TO NEW
    entries = await db.clock_entries.find(query, {"_id": 0}).sort("clock_in_time", 1).to_list(10000)
    
    # Calculate totals
    total_hours = 0
    hours_per_project = {}
    
    processed_entries = []
    for entry in entries:
        # Parse dates
        if isinstance(entry.get('created_at'), str):
            entry['created_at'] = datetime.fromisoformat(entry['created_at'])
        if isinstance(entry.get('clock_in_time'), str):
            entry['clock_in_time'] = datetime.fromisoformat(entry['clock_in_time'])
        if entry.get('clock_out_time') and isinstance(entry['clock_out_time'], str):
            entry['clock_out_time'] = datetime.fromisoformat(entry['clock_out_time'])
        
        # Calculate hours
        hours = entry.get('total_hours', 0) or 0
        total_hours += hours
        
        # Hours per project
        project_name = entry.get('project_name', 'Unknown')
        if project_name not in hours_per_project:
            hours_per_project[project_name] = 0
        hours_per_project[project_name] += hours
        
        processed_entries.append(entry)
    
    return {
        "entries": processed_entries,
        "total_hours": round(total_hours, 2),
        "hours_per_project": hours_per_project,
        "entry_count": len(processed_entries)
    }

@api_router.get("/clock/entries/export/pdf")
async def export_my_entries_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    # Get user's entries
    query = {"user_id": current_user.id, "status": "clocked_out"}
    
    entries = await db.clock_entries.find(query, {"_id": 0}).sort("clock_in_time", -1).to_list(10000)
    
    # Filter by date if needed
    if start_date or end_date:
        filtered_entries = []
        for entry in entries:
            entry_date = entry["clock_in_time"][:10] if isinstance(entry["clock_in_time"], str) else entry["clock_in_time"].strftime("%Y-%m-%d")
            if start_date and entry_date < start_date:
                continue
            if end_date and entry_date > end_date:
                continue
            filtered_entries.append(entry)
        entries = filtered_entries
    
    # Create PDF
    pdf_file = io.BytesIO()
    doc = SimpleDocTemplate(pdf_file, pagesize=A4)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"<b>Mijn Urenregistratie - {current_user.full_name}</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    if not entries:
        no_data = Paragraph("Geen voltooide registraties gevonden.", styles['Normal'])
        elements.append(no_data)
    else:
        # Create table
        table_data = [['Datum', 'Bedrijf', 'Project', 'In', 'Uit', 'Uren']]
        
        total_hours = 0
        for entry in entries:
            clock_in_time = entry["clock_in_time"][:16] if isinstance(entry["clock_in_time"], str) else entry["clock_in_time"].strftime("%H:%M")
            clock_out_time = entry["clock_out_time"][:16] if isinstance(entry.get("clock_out_time"), str) else entry.get("clock_out_time", datetime.now()).strftime("%H:%M")
            entry_date = entry["clock_in_time"][:10] if isinstance(entry["clock_in_time"], str) else entry["clock_in_time"].strftime("%Y-%m-%d")
            
            table_data.append([
                entry_date,
                entry["company"][:15],
                entry["project_name"][:20],
                clock_in_time,
                clock_out_time,
                str(entry.get("total_hours", 0))
            ])
            total_hours += entry.get("total_hours", 0)
        
        # Add total
        table_data.append(['', '', '', '', 'Totaal', str(round(total_hours, 2))])
        
        table = Table(table_data, colWidths=[2.5*cm, 3.5*cm, 4*cm, 2*cm, 2*cm, 2*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16a085')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
    
    doc.build(elements)
    pdf_file.seek(0)
    
    filename = f"mijn_uren_{current_user.first_name}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        pdf_file,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Reports endpoints
@api_router.get("/reports/excel")
async def export_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_id: Optional[str] = None,
    admin: User = Depends(get_admin_user)
):
    query = {"status": "clocked_out"}
    if user_id:
        query["user_id"] = user_id
    
    entries = await db.clock_entries.find(query, {"_id": 0}).sort("clock_in_time", -1).to_list(10000)
    
    # Filter by date if needed
    if start_date or end_date:
        filtered_entries = []
        for entry in entries:
            entry_date = entry["clock_in_time"][:10] if isinstance(entry["clock_in_time"], str) else entry["clock_in_time"].strftime("%Y-%m-%d")
            if start_date and entry_date < start_date:
                continue
            if end_date and entry_date > end_date:
                continue
            filtered_entries.append(entry)
        entries = filtered_entries
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Urenregistratie"
    
    header_fill = PatternFill(start_color="16a085", end_color="16a085", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ["Datum", "Medewerker", "Bedrijf", "Project", "Locatie", "Ingeklokt", "Uitgeklokt", "Totaal Uren", "Opmerking"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, entry in enumerate(entries, 2):
        clock_in = entry["clock_in_time"][:10] if isinstance(entry["clock_in_time"], str) else entry["clock_in_time"].strftime("%Y-%m-%d")
        clock_in_time = entry["clock_in_time"][11:16] if isinstance(entry["clock_in_time"], str) else entry["clock_in_time"].strftime("%H:%M")
        clock_out_time = entry["clock_out_time"][11:16] if isinstance(entry.get("clock_out_time"), str) else entry.get("clock_out_time", datetime.now()).strftime("%H:%M")
        
        ws.cell(row=row_idx, column=1, value=clock_in)
        ws.cell(row=row_idx, column=2, value=entry["user_name"])
        ws.cell(row=row_idx, column=3, value=entry["company"])
        ws.cell(row=row_idx, column=4, value=entry["project_name"])
        ws.cell(row=row_idx, column=5, value=entry["project_location"])
        ws.cell(row=row_idx, column=6, value=clock_in_time)
        ws.cell(row=row_idx, column=7, value=clock_out_time)
        ws.cell(row=row_idx, column=8, value=entry.get("total_hours", 0))
        ws.cell(row=row_idx, column=9, value=entry.get("note", ""))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 40
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"urenregistratie_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Users endpoint
@api_router.get("/users", response_model=List[User])
async def get_users(admin: User = Depends(get_admin_user)):
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    for user in users:
        if isinstance(user['created_at'], str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
    return users

@api_router.put("/users/{user_id}", response_model=User)
async def update_user(user_id: str, user_data: UserUpdate, admin: User = Depends(get_admin_user)):
    update_dict = {k: v for k, v in user_data.model_dump().items() if v is not None}
    
    if not update_dict:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": update_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    if isinstance(user['created_at'], str):
        user['created_at'] = datetime.fromisoformat(user['created_at'])
    return User(**user)

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: User = Depends(get_admin_user)):
    # Don't allow deleting yourself
    if user_id == admin.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Also delete all clock entries for this user
    await db.clock_entries.delete_many({"user_id": user_id})
    
    return {"message": "User deleted successfully"}

# Initialize admin
@api_router.post("/init-admin")
async def init_admin():
    admin_exists = await db.users.find_one({"role": "admin"})
    if admin_exists:
        raise HTTPException(status_code=400, detail="Admin already exists")
    
    admin = User(
        email="admin@theglobal.nl",
        first_name="Admin",
        last_name="Administrator",
        role="admin"
    )
    admin_dict = admin.model_dump()
    admin_dict["password"] = get_password_hash("admin123")
    admin_dict["created_at"] = admin_dict["created_at"].isoformat()
    
    await db.users.insert_one(admin_dict)
    return {"message": "Admin created", "email": "admin@theglobal.nl", "password": "admin123"}

# Admin PDF Export with filters
@api_router.get("/admin/export/pdf")
async def export_admin_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    project_ids: Optional[str] = None,
    user_ids: Optional[str] = None,
    admin: User = Depends(get_admin_user)
):
    """Export admin overview to PDF with filters"""
    # Build query
    query = {"status": "clocked_out"}
    
    if start_date or end_date:
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        if date_query:
            query["$expr"] = {
                "$and": [
                    {"$gte": [{"$substr": ["$clock_in_time", 0, 10]}, date_query.get("$gte", "1900-01-01")]},
                    {"$lte": [{"$substr": ["$clock_in_time", 0, 10]}, date_query.get("$lte", "9999-12-31")]}
                ]
            }
    
    if project_ids:
        query["project_id"] = {"$in": project_ids.split(",")}
    
    if user_ids:
        query["user_id"] = {"$in": user_ids.split(",")}
    
    entries = await db.clock_entries.find(query, {"_id": 0}).sort("clock_in_time", -1).to_list(10000)
    
    # Calculate totals
    total_hours = sum(entry.get("total_hours", 0) or 0 for entry in entries)
    
    # Create PDF
    pdf_file = io.BytesIO()
    doc = SimpleDocTemplate(pdf_file, pagesize=A4, leftMargin=1*cm, rightMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Header
    title = Paragraph("<b>The Global Bedrijfsdiensten BV</b>", styles['Title'])
    elements.append(title)
    subtitle = Paragraph(f"<b>Urenoverzicht</b><br/>Periode: {start_date or 'Alle'} t/m {end_date or 'Alle'}", styles['Heading2'])
    elements.append(subtitle)
    elements.append(Spacer(1, 0.5*cm))
    
    # Summary
    summary_text = f"<b>Totaal uren:</b> {total_hours:.2f}u | <b>Aantal registraties:</b> {len(entries)}"
    summary = Paragraph(summary_text, styles['Normal'])
    elements.append(summary)
    elements.append(Spacer(1, 0.5*cm))
    
    if not entries:
        no_data = Paragraph("Geen voltooide registraties gevonden.", styles['Normal'])
        elements.append(no_data)
    else:
        # Table with all required columns
        table_data = [['Datum', 'Medewerker', 'Project', 'Start', 'Eind', 'Uren', 'Locatie', 'Afstand (m)', 'Match', 'Opmerking']]
        
        for entry in entries:
            entry_date = entry["clock_in_time"][:10] if isinstance(entry["clock_in_time"], str) else entry["clock_in_time"].strftime("%Y-%m-%d")
            start_time = entry["clock_in_time"][11:16] if isinstance(entry["clock_in_time"], str) else entry["clock_in_time"].strftime("%H:%M")
            end_time = entry.get("clock_out_time", "")
            if end_time:
                end_time = end_time[11:16] if isinstance(end_time, str) else end_time.strftime("%H:%M")
            else:
                end_time = "-"
            
            loc_lat = entry.get("clock_in_location", {}).get("latitude", 0)
            loc_lon = entry.get("clock_in_location", {}).get("longitude", 0)
            location_str = f"{loc_lat:.4f}, {loc_lon:.4f}" if loc_lat else "-"
            
            distance = entry.get("distance_to_project_m")
            distance_str = f"{int(distance)}" if distance is not None else "-"
            
            project_match = entry.get("project_match")
            match_str = "JA" if project_match is True else ("NEE" if project_match is False else "-")
            
            table_data.append([
                entry_date,
                entry.get("user_name", "")[:15],
                entry.get("project_name", "")[:12],
                start_time,
                end_time,
                f"{entry.get('total_hours', 0):.1f}",
                location_str[:15],
                distance_str,
                match_str,
                (entry.get("note", "") or "")[:10]
            ])
        
        table = Table(table_data, colWidths=[1.8*cm, 2.2*cm, 2.2*cm, 1.2*cm, 1.2*cm, 1.2*cm, 2.5*cm, 1.3*cm, 1*cm, 2*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16a085')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ]))
        
        elements.append(table)
    
    doc.build(elements)
    pdf_file.seek(0)
    
    filename = f"urenoverzicht_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.pdf"
    
    return StreamingResponse(
        pdf_file,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Mandagenstaat endpoint
@api_router.get("/admin/mandagenstaat")
async def get_mandagenstaat_data(
    start_date: str,
    end_date: str,
    project_id: str,
    user_id: Optional[str] = None,
    admin: User = Depends(get_admin_user)
):
    """Get mandagenstaat data for specified period and project"""
    # Get project details
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Build query
    query = {
        "status": "clocked_out",
        "project_id": project_id,
        "$expr": {
            "$and": [
                {"$gte": [{"$substr": ["$clock_in_time", 0, 10]}, start_date]},
                {"$lte": [{"$substr": ["$clock_in_time", 0, 10]}, end_date]}
            ]
        }
    }
    
    if user_id:
        query["user_id"] = user_id
    
    entries = await db.clock_entries.find(query, {"_id": 0}).sort("clock_in_time", 1).to_list(10000)
    
    # Group by date and user
    grouped_data = {}
    user_totals = {}
    
    for entry in entries:
        entry_date = entry["clock_in_time"][:10] if isinstance(entry["clock_in_time"], str) else entry["clock_in_time"].strftime("%Y-%m-%d")
        user_name = entry.get("user_name", "Unknown")
        hours = entry.get("total_hours", 0) or 0
        
        if entry_date not in grouped_data:
            grouped_data[entry_date] = {}
        
        if user_name not in grouped_data[entry_date]:
            grouped_data[entry_date][user_name] = {
                "hours": 0,
                "notes": []
            }
        
        grouped_data[entry_date][user_name]["hours"] += hours
        if entry.get("note"):
            grouped_data[entry_date][user_name]["notes"].append(entry["note"])
        
        if user_name not in user_totals:
            user_totals[user_name] = 0
        user_totals[user_name] += hours
    
    return {
        "project": project,
        "start_date": start_date,
        "end_date": end_date,
        "grouped_data": grouped_data,
        "user_totals": user_totals,
        "total_hours": sum(user_totals.values())
    }
@api_router.get("/admin/mandagenstaat/export/excel")
async def export_mandagenstaat_excel(
    start_date: str,
    end_date: str,
    project_id: str,
    user_id: Optional[str] = None,
    admin: User = Depends(get_admin_user)
):
    """Export mandagenstaat to Excel - TEMPLATE-BASED (gebruikt user template)"""
    from mandagenstaat_template_based import create_from_template
    
    # Get project
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Build query
    query = {
        "status": "clocked_out",
        "project_id": project_id,
        "$expr": {
            "$and": [
                {"$gte": [{"$substr": ["$clock_in_time", 0, 10]}, start_date]},
                {"$lte": [{"$substr": ["$clock_in_time", 0, 10]}, end_date]}
            ]
        }
    }
    
    if user_id and user_id != "all":
        query["user_id"] = user_id
    
    entries = await db.clock_entries.find(query, {"_id": 0}).sort("clock_in_time", 1).to_list(10000)
    
    # Group by user and weekday
    user_week_data = {}
    
    for entry in entries:
        user_name = entry.get("user_name", "Unknown")
        user_id_val = entry.get("user_id", "")
        entry_datetime = entry["clock_in_time"] if isinstance(entry["clock_in_time"], datetime) else datetime.fromisoformat(entry["clock_in_time"])
        weekday = entry_datetime.weekday()  # 0=Monday, 6=Sunday
        hours = entry.get("total_hours", 0) or 0
        
        if user_name not in user_week_data:
            user_week_data[user_name] = {
                "user_id": user_id_val,
                "days": [0, 0, 0, 0, 0, 0, 0]
            }
        
        user_week_data[user_name]["days"][weekday] += hours
    
    # Get BSN for users
    for user_name in user_week_data:
        user_id_val = user_week_data[user_name]["user_id"]
        user_doc = await db.users.find_one({"id": user_id_val}, {"_id": 0, "bsn": 1})
        user_week_data[user_name]["bsn"] = user_doc.get("bsn", "") if user_doc else ""
    
    # Create Excel from USER TEMPLATE
    excel_file = create_from_template(project, user_week_data, start_date, end_date)
    
    # Filename met runtime datum: Mandagenstaat_dd-mm-yyyy_Bedrijfsnaam.xlsx
    runtime_date = datetime.now().strftime("%d-%m-%Y")  # Runtime datum
    company_name = project.get('company', 'Bedrijf').replace(' ', '_')
    filename = f"Mandagenstaat_{runtime_date}_{company_name}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/admin/mandagenstaat/export/pdf")
async def export_mandagenstaat_pdf(
    start_date: str,
    end_date: str,
    project_id: str,
    user_id: Optional[str] = None,
    admin: User = Depends(get_admin_user)
):
    """Export mandagenstaat to PDF - SSCONVERT (Gnumeric, geen watermark)"""
    import logging
    import traceback
    import subprocess
    import tempfile
    import os
    import io
    
    logger = logging.getLogger(__name__)
    
    try:
        from mandagenstaat_template_based import create_from_template
        
        # Get project
        project = await db.projects.find_one({"id": project_id}, {"_id": 0})
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        
        # Build query
        query = {
            "status": "clocked_out",
            "project_id": project_id,
            "$expr": {
                "$and": [
                    {"$gte": [{"$substr": ["$clock_in_time", 0, 10]}, start_date]},
                    {"$lte": [{"$substr": ["$clock_in_time", 0, 10]}, end_date]}
                ]
            }
        }
        
        if user_id and user_id != "all":
            query["user_id"] = user_id
        
        entries = await db.clock_entries.find(query, {"_id": 0}).sort("clock_in_time", 1).to_list(10000)
        
        # Group by user and weekday
        user_week_data = {}
        
        for entry in entries:
            user_name = entry.get("user_name", "Unknown")
            user_id_val = entry.get("user_id", "")
            entry_datetime = entry["clock_in_time"] if isinstance(entry["clock_in_time"], datetime) else datetime.fromisoformat(entry["clock_in_time"])
            weekday = entry_datetime.weekday()
            hours = entry.get("total_hours", 0) or 0
            
            if user_name not in user_week_data:
                user_week_data[user_name] = {
                    "user_id": user_id_val,
                    "days": [0, 0, 0, 0, 0, 0, 0]
                }
            
            user_week_data[user_name]["days"][weekday] += hours
        
        # Get BSN for users
        for user_name in user_week_data:
            user_id_val = user_week_data[user_name]["user_id"]
            user_doc = await db.users.find_one({"id": user_id_val}, {"_id": 0, "bsn": 1})
            user_week_data[user_name]["bsn"] = user_doc.get("bsn", "") if user_doc else ""
    
        # Create Excel met correcte print settings
        excel_file = create_from_template(project, user_week_data, start_date, end_date)
        
        # Save Excel to temp
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', mode='wb') as tmp_excel:
            tmp_excel.write(excel_file.read())
            excel_path = tmp_excel.name
        
        # Output PDF path
        pdf_path = excel_path.replace('.xlsx', '.pdf')
        
        try:
            # CHECK: Ensure ssconvert is available
            import shutil
            if not shutil.which('ssconvert'):
                # Don't try to install on cloud platforms - use alternative PDF generation
                logger.warning("ssconvert not found, using alternative PDF generation method...")
                # Use template-based PDF generation which doesn't require ssconvert
                pdf_data = create_pdf_from_template(project, user_week_data, start_date, end_date)
            else:
                # SSCONVERT: Excel → PDF (respecteert Excel print settings)
                result = subprocess.run([
                    'ssconvert',
                    excel_path,
                    pdf_path,
                    '--export-type=Gnumeric_pdf:pdf_assistant'
                ], timeout=30, capture_output=True, text=True)
                
                if result.returncode != 0:
                    logger.error(f"ssconvert stderr: {result.stderr}")
                    # Fallback to template-based PDF generation
                    logger.warning("ssconvert failed, falling back to alternative PDF generation...")
                    pdf_data = create_pdf_from_template(project, user_week_data, start_date, end_date)
                else:
                    if not os.path.exists(pdf_path):
                        raise Exception(f"PDF not created at: {pdf_path}")
                    
                    # NOTE: ssconvert neemt automatisch images mee uit Excel template
                    # Geen extra logo insert nodig
                    
                    # Read PDF
                    with open(pdf_path, 'rb') as f:
                        pdf_data = io.BytesIO(f.read())
            
        finally:
            # Cleanup
            try:
                if os.path.exists(excel_path):
                    os.unlink(excel_path)
                if os.path.exists(pdf_path):
                    os.unlink(pdf_path)
            except:
                pass
        
        # Filename
        runtime_date = datetime.now().strftime("%d-%m-%Y")
        company_name = project.get('company', 'Bedrijf').replace(' ', '_')
        filename = f"Mandagenstaat_{runtime_date}_{company_name}.pdf"
        
        return StreamingResponse(
            pdf_data,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Type": "application/pdf",
                "Cache-Control": "no-cache"
            }
        )
    
    except Exception as e:
        logger.error(f"PDF generation failed: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"PDF kon niet worden gegenereerd: {str(e)}"
        )


@api_router.get("/admin/backup/export")
async def export_backup(admin: User = Depends(get_admin_user)):
    """Export complete database backup as JSON"""
    import json
    
    backup_data = {
        "backup_date": datetime.now(timezone.utc).isoformat(),
        "version": "1.0",
        "data": {}
    }
    
    # Export users (without passwords)
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(10000)
    for user in users:
        if isinstance(user.get('created_at'), datetime):
            user['created_at'] = user['created_at'].isoformat()
    backup_data["data"]["users"] = users
    
    # Export projects
    projects = await db.projects.find({}, {"_id": 0}).to_list(10000)
    for project in projects:
        if isinstance(project.get('created_at'), datetime):
            project['created_at'] = project['created_at'].isoformat()
    backup_data["data"]["projects"] = projects
    
    # Export clock entries
    clock_entries = await db.clock_entries.find({}, {"_id": 0}).to_list(100000)
    for entry in clock_entries:
        if isinstance(entry.get('clock_in_time'), datetime):
            entry['clock_in_time'] = entry['clock_in_time'].isoformat()
        if isinstance(entry.get('clock_out_time'), datetime):
            entry['clock_out_time'] = entry['clock_out_time'].isoformat()
    backup_data["data"]["clock_entries"] = clock_entries
    
    # Export invitations
    invitations = await db.invitations.find({}, {"_id": 0}).to_list(10000)
    for inv in invitations:
        if isinstance(inv.get('created_at'), datetime):
            inv['created_at'] = inv['created_at'].isoformat()
        if isinstance(inv.get('expires_at'), datetime):
            inv['expires_at'] = inv['expires_at'].isoformat()
    backup_data["data"]["invitations"] = invitations
    
    # Convert to JSON
    json_data = json.dumps(backup_data, indent=2, ensure_ascii=False)
    json_bytes = io.BytesIO(json_data.encode('utf-8'))
    
    filename = f"urenregistratie_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    
    return StreamingResponse(
        json_bytes,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.post("/admin/backup/import")
async def import_backup(
    file: UploadFile = File(...),
    admin: User = Depends(get_admin_user)
):
    """Import database backup from JSON file"""
    import json
    
    try:
        # Read uploaded file
        content = await file.read()
        backup_data = json.loads(content.decode('utf-8'))
        
        if "data" not in backup_data:
            raise HTTPException(status_code=400, detail="Invalid backup file format")
        
        data = backup_data["data"]
        imported_counts = {
            "users": 0,
            "projects": 0,
            "clock_entries": 0,
            "invitations": 0
        }
        
        # Import users (skip if email already exists)
        if "users" in data:
            for user in data["users"]:
                # Convert date strings back to datetime
                if 'created_at' in user and isinstance(user['created_at'], str):
                    user['created_at'] = datetime.fromisoformat(user['created_at'])
                
                # Check if user exists
                existing = await db.users.find_one({"email": user["email"]}, {"_id": 0})
                if not existing:
                    # Set default password for imported users without one
                    if 'password' not in user:
                        user['password'] = bcrypt.hashpw("changeme123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                    await db.users.insert_one(user)
                    imported_counts["users"] += 1
        
        # Import projects (skip if id already exists)
        if "projects" in data:
            for project in data["projects"]:
                if 'created_at' in project and isinstance(project['created_at'], str):
                    project['created_at'] = datetime.fromisoformat(project['created_at'])
                
                existing = await db.projects.find_one({"id": project["id"]}, {"_id": 0})
                if not existing:
                    await db.projects.insert_one(project)
                    imported_counts["projects"] += 1
        
        # Import clock entries (skip if id already exists)
        if "clock_entries" in data:
            for entry in data["clock_entries"]:
                if 'clock_in_time' in entry and isinstance(entry['clock_in_time'], str):
                    entry['clock_in_time'] = datetime.fromisoformat(entry['clock_in_time'])
                if 'clock_out_time' in entry and isinstance(entry['clock_out_time'], str):
                    entry['clock_out_time'] = datetime.fromisoformat(entry['clock_out_time'])
                
                existing = await db.clock_entries.find_one({"id": entry["id"]}, {"_id": 0})
                if not existing:
                    await db.clock_entries.insert_one(entry)
                    imported_counts["clock_entries"] += 1
        
        # Import invitations (skip if token already exists)
        if "invitations" in data:
            for inv in data["invitations"]:
                if 'created_at' in inv and isinstance(inv['created_at'], str):
                    inv['created_at'] = datetime.fromisoformat(inv['created_at'])
                if 'expires_at' in inv and isinstance(inv['expires_at'], str):
                    inv['expires_at'] = datetime.fromisoformat(inv['expires_at'])
                
                existing = await db.invitations.find_one({"token": inv["token"]}, {"_id": 0})
                if not existing:
                    await db.invitations.insert_one(inv)
                    imported_counts["invitations"] += 1
        
        return {
            "message": "Backup imported successfully",
            "imported": imported_counts,
            "backup_date": backup_data.get("backup_date", "unknown")
        }
        
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON file")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
